"""
RuralCaixa — importacao_financeiro.py
Endpoint de importacao em lote para o Livro Caixa (tela "Financeiro").

Aceita .csv ou .xlsx com as colunas:
    data, tipo, categoria, descricao, valor, documento (opcional), observacoes (opcional)

- data: formato DD/MM/AAAA ou AAAA-MM-DD
- tipo: "receita" ou "despesa"
- categoria: texto livre (ex: venda_producao, insumos, mao_de_obra...)
- valor: numero (aceita virgula ou ponto decimal)

Adicione em app/main.py, perto do router livro_caixa:
    from app.routers.importacao_financeiro import router as importacao_financeiro_router
    if importacao_financeiro_router: app.include_router(importacao_financeiro_router)

Requer 'openpyxl' no requirements.txt para .xlsx (adicionar se ainda nao tiver).
"""
from fastapi import APIRouter, HTTPException, UploadFile, File, Query
from typing import Optional
from datetime import datetime, date
import os
import io
import csv
import re
import psycopg2
import psycopg2.extras

router = APIRouter(prefix="/livro-caixa", tags=["Livro Caixa - Importacao"])

DB_URL = os.getenv("DATABASE_URL", "")


def get_db():
    return psycopg2.connect(DB_URL, cursor_factory=psycopg2.extras.RealDictCursor)


COLUNAS_ESPERADAS = ["data", "tipo", "categoria", "descricao", "valor"]


def normalizar_cabecalho(nome: str) -> str:
    nome = nome.strip().lower()
    nome = re.sub(r"[áàâã]", "a", nome)
    nome = re.sub(r"[éê]", "e", nome)
    nome = re.sub(r"[íî]", "i", nome)
    nome = re.sub(r"[óôõ]", "o", nome)
    nome = re.sub(r"[úû]", "u", nome)
    nome = re.sub(r"[^a-z0-9_]", "_", nome)
    return nome


def parse_data(valor: str) -> Optional[date]:
    valor = (valor or "").strip()
    if not valor:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(valor, fmt).date()
        except ValueError:
            continue
    return None


def parse_valor(valor) -> Optional[float]:
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    valor = str(valor).strip()
    if not valor:
        return None
    valor = valor.replace("R$", "").replace(" ", "")
    # Trata formato brasileiro 1.234,56 e tambem 1234.56
    if "," in valor and "." in valor:
        valor = valor.replace(".", "").replace(",", ".")
    elif "," in valor:
        valor = valor.replace(",", ".")
    try:
        return float(valor)
    except ValueError:
        return None


def ler_linhas_csv(conteudo: bytes) -> list[dict]:
    texto = conteudo.decode("utf-8-sig", errors="replace")
    leitor = csv.reader(io.StringIO(texto), delimiter=";" if ";" in texto.splitlines()[0] else ",")
    linhas = list(leitor)
    if not linhas:
        return []
    cabecalho = [normalizar_cabecalho(c) for c in linhas[0]]
    resultado = []
    for linha in linhas[1:]:
        if not any(c.strip() for c in linha):
            continue
        resultado.append(dict(zip(cabecalho, linha)))
    return resultado


def ler_linhas_xlsx(conteudo: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl nao instalado no servidor. Adicione 'openpyxl' ao requirements.txt e refaça o deploy.",
        )
    wb = load_workbook(io.BytesIO(conteudo), data_only=True)
    ws = wb.active
    linhas_raw = list(ws.iter_rows(values_only=True))
    if not linhas_raw:
        return []
    cabecalho = [normalizar_cabecalho(str(c or "")) for c in linhas_raw[0]]
    resultado = []
    for linha in linhas_raw[1:]:
        if not any(c is not None and str(c).strip() for c in linha):
            continue
        valores = [("" if c is None else c) for c in linha]
        resultado.append(dict(zip(cabecalho, valores)))
    return resultado


@router.post("/{imovel_id}/importar")
async def importar_lancamentos(
    imovel_id: int,
    arquivo: UploadFile = File(...),
    ano_base: Optional[int] = Query(None),
):
    conteudo = await arquivo.read()
    nome = (arquivo.filename or "").lower()

    if nome.endswith(".xlsx") or nome.endswith(".xls"):
        linhas = ler_linhas_xlsx(conteudo)
    elif nome.endswith(".csv"):
        linhas = ler_linhas_csv(conteudo)
    else:
        raise HTTPException(status_code=400, detail="Formato não suportado. Envie .csv ou .xlsx")

    faltando = [c for c in COLUNAS_ESPERADAS if not any(c in l for l in linhas[:1])]
    if linhas and faltando:
        raise HTTPException(
            status_code=400,
            detail=f"Colunas obrigatórias ausentes: {', '.join(faltando)}. "
                   f"Esperado: {', '.join(COLUNAS_ESPERADAS)}",
        )

    conn = get_db()
    cur = conn.cursor()
    importados = 0
    erros = []

    for i, linha in enumerate(linhas, start=2):  # linha 1 = cabecalho
        try:
            data_lanc = parse_data(str(linha.get("data", "")))
            tipo = str(linha.get("tipo", "")).strip().lower()
            categoria = str(linha.get("categoria", "")).strip() or "outras_despesas"
            descricao = str(linha.get("descricao", "")).strip()
            valor = parse_valor(linha.get("valor"))
            documento = str(linha.get("documento", "")).strip() or None
            observacoes = str(linha.get("observacoes", "")).strip() or None

            if not data_lanc:
                erros.append({"linha": i, "msg": "Data inválida ou ausente"})
                continue
            if tipo not in ("receita", "despesa"):
                erros.append({"linha": i, "msg": f"Tipo inválido: '{tipo}' (use receita ou despesa)"})
                continue
            if valor is None or valor <= 0:
                erros.append({"linha": i, "msg": "Valor inválido ou ausente"})
                continue
            if not descricao:
                erros.append({"linha": i, "msg": "Descrição ausente"})
                continue

            ano = ano_base or data_lanc.year
            cur.execute("""
                INSERT INTO livro_caixa_lancamentos
                    (imovel_id, ano_base, data_lancamento, tipo, categoria, descricao,
                     valor, origem, deducao_irpf, documento, observacoes)
                VALUES (%s,%s,%s,%s,%s,%s,%s,'importacao',true,%s,%s)
            """, (imovel_id, ano, data_lanc, tipo, categoria, descricao, valor,
                  documento, observacoes))
            importados += 1
        except Exception as e:
            erros.append({"linha": i, "msg": str(e)})

    conn.commit()
    conn.close()

    return {
        "total_linhas": len(linhas),
        "importados": importados,
        "erros": erros,
    }
