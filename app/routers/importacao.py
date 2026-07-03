"""
RuralCaixa — app/routers/importacao.py
Router de importacao em lote — restaura o endpoint POST /importacao/lancamentos
que a procedure tRPC `importarLancamentos` (server/routers/railway.ts) ja
chama, mas que estava ausente desde a reestruturacao de 243 arquivos.

Contrato esperado pelo frontend (multipart/form-data):
    arquivo         : arquivo .csv ou .xlsx
    produtor_id     : int
    imovel_id       : int (recebido, mas atualmente nao usado no INSERT —
                      LancamentoCreate/criar_lancamento nao tem esse campo)
    mapa_data       : nome da coluna que contem a data (opcional, default "data")
    mapa_valor      : nome da coluna que contem o valor (opcional, default "valor")
    mapa_descricao  : nome da coluna que contem a descricao (opcional, default "descricao")
    mapa_tipo       : nome da coluna que contem o tipo receita/despesa (opcional, default "tipo")

Resposta:
    { criados: int, erros: int, total: int, mensagem?: str }

Reaproveita a MESMA logica de classificacao automatica (regras_classificacao +
classificar()) que o endpoint manual POST /lancamentos ja usa, chamando
diretamente a funcao criar_lancamento() de app.main — import feito dentro da
funcao para evitar import circular (main.py importa este router).
"""
from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from typing import Optional
from datetime import datetime, date
import io
import csv
import re

router = APIRouter(prefix="/importacao", tags=["Importacao Generica"])

# Sinônimos aceitos para cada campo, quando o usuário não informa mapeamento manual.
# Ordem importa: primeiro sinônimo encontrado no cabeçalho vence.
SINONIMOS = {
    "data": ["data", "data_transacao", "dt_lancamento", "data_lancamento", "dt"],
    "valor": ["valor", "valor_rs", "valor_r", "vl_total", "montante"],
    "descricao": ["descricao", "historico", "descr", "obs", "observacao"],
    "tipo": ["tipo", "natureza", "tipo_lancamento"],
    "status": ["status", "situacao"],
    # Colunas extras usadas como fallback quando "descricao" vem vazia/"—"
    "categoria_fallback": ["natureza", "plano_de_contas", "categoria"],
}

# Valores de status que indicam que a transação JÁ aconteceu de verdade.
# Qualquer outro valor reconhecido (ex: "Planejamento", "Plan. Obrigatório",
# "Previsto", "Pendente") é tratado como projeção/futuro e NÃO é importado —
# livro caixa/LCDPR deve refletir movimentação de caixa real, não planejada.
# Se a planilha não tiver coluna de status, importa tudo normalmente (esse
# filtro só entra em ação quando a coluna existe).
STATUS_REALIZADOS = {
    "realizada", "realizado", "concluida", "concluido", "efetivada", "efetivado",
    "confirmada", "confirmado", "pago", "paga", "liquidada", "liquidado",
}

PALAVRAS_CABECALHO = ["data", "valor", "descri", "historico", "natureza", "status", "tipo"]


def normalizar_cabecalho(nome: str) -> str:
    nome = nome.strip().lower()
    nome = re.sub(r"[áàâã]", "a", nome)
    nome = re.sub(r"[éê]", "e", nome)
    nome = re.sub(r"[íî]", "i", nome)
    nome = re.sub(r"[óôõ]", "o", nome)
    nome = re.sub(r"[úû]", "u", nome)
    nome = re.sub(r"[ç]", "c", nome)
    nome = re.sub(r"[^a-z0-9_]", "_", nome)
    nome = re.sub(r"_+", "_", nome).strip("_")
    return nome


def parse_data(valor) -> Optional[str]:
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    if isinstance(valor, date):
        return valor.isoformat()
    valor = str(valor).strip()
    if not valor:
        return None
    valor = valor.split(" ")[0]
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(valor, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def parse_valor(valor) -> Optional[float]:
    """Aceita numero puro, ou texto no formato brasileiro/monetario:
    '-R$  250,00', 'R$ 1.234,56', '250,00', '250.00', '-250'..."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip()
    if not texto or texto == "—":
        return None
    negativo = "-" in texto
    texto = texto.replace("R$", "").replace("-", "").replace(" ", "").strip()
    if not texto:
        return None
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "," in texto:
        texto = texto.replace(",", ".")
    try:
        num = float(texto)
        return -num if negativo else num
    except ValueError:
        return None


def normalizar_valor_status(valor: str) -> str:
    """Normaliza um VALOR de status (não um nome de coluna) pra comparar com
    STATUS_REALIZADOS — minúsculo, sem acento, sem pontuação."""
    texto = str(valor or "").strip().lower()
    texto = re.sub(r"[áàâã]", "a", texto)
    texto = re.sub(r"[éê]", "e", texto)
    texto = re.sub(r"[íî]", "i", texto)
    texto = re.sub(r"[óôõ]", "o", texto)
    texto = re.sub(r"[úû]", "u", texto)
    texto = re.sub(r"[ç]", "c", texto)
    texto = re.sub(r"[^a-z]", "", texto)
    return texto


def linha_parece_cabecalho(celulas: list) -> bool:
    """Detecta se uma linha e o cabecalho real, contando quantas palavras-chave
    esperadas aparecem nela (ignora linhas de metadados tipo 'Gerado em...')."""
    textos = [normalizar_cabecalho(str(c or "")) for c in celulas]
    acertos = sum(1 for t in textos for palavra in PALAVRAS_CABECALHO if palavra in t)
    return acertos >= 2


def resolver_coluna(cabecalho_normalizado: list, campo: str, mapa_manual: Optional[str]) -> Optional[str]:
    """Decide qual coluna do cabecalho usar para um campo (data/valor/descricao/tipo).
    Prioridade: mapeamento manual do usuario > sinonimos conhecidos."""
    if mapa_manual:
        col = normalizar_cabecalho(mapa_manual)
        if col in cabecalho_normalizado:
            return col
    for sinonimo in SINONIMOS.get(campo, []):
        if sinonimo in cabecalho_normalizado:
            return sinonimo
    return None


def ler_linhas_csv(conteudo: bytes) -> tuple[list[dict], list[str]]:
    texto = conteudo.decode("utf-8-sig", errors="replace")
    linhas_texto = texto.splitlines()
    primeira_linha = linhas_texto[0] if linhas_texto else ""
    leitor = csv.reader(io.StringIO(texto), delimiter=";" if ";" in primeira_linha else ",")
    todas_linhas = list(leitor)
    return _processar_linhas_com_deteccao(todas_linhas)


def ler_linhas_xlsx(conteudo: bytes) -> tuple[list[dict], list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl nao instalado no servidor. Adicione 'openpyxl' ao requirements.txt.",
        )
    wb = load_workbook(io.BytesIO(conteudo), data_only=True)
    ws = wb.active
    linhas_raw = list(ws.iter_rows(values_only=True))
    todas_linhas = [[("" if c is None else c) for c in linha] for linha in linhas_raw]
    return _processar_linhas_com_deteccao(todas_linhas)


def _processar_linhas_com_deteccao(todas_linhas: list) -> tuple[list[dict], list[str]]:
    """Procura a linha de cabecalho real nas primeiras 20 linhas (pula linhas
    de metadados tipo 'Gerado em...', 'Fazenda...', 'Total de...'), depois
    monta os dicionarios de dados a partir dela em diante."""
    if not todas_linhas:
        return [], []

    idx_cabecalho = 0
    for i, linha in enumerate(todas_linhas[:20]):
        if linha_parece_cabecalho(linha):
            idx_cabecalho = i
            break

    cabecalho_bruto = [str(c or "") for c in todas_linhas[idx_cabecalho]]
    cabecalho = [normalizar_cabecalho(c) for c in cabecalho_bruto]

    resultado = []
    for linha in todas_linhas[idx_cabecalho + 1:]:
        if not any(str(c).strip() for c in linha):
            continue
        resultado.append(dict(zip(cabecalho, linha)))
    return resultado, cabecalho


@router.post("/lancamentos")
async def importar_lancamentos(
    arquivo: UploadFile = File(...),
    produtor_id: int = Form(...),
    imovel_id: Optional[int] = Form(None),
    mapa_data: Optional[str] = Form(None),
    mapa_valor: Optional[str] = Form(None),
    mapa_descricao: Optional[str] = Form(None),
    mapa_tipo: Optional[str] = Form(None),
):
    # Import tardio para evitar dependencia circular com app.main
    from app.main import criar_lancamento, LancamentoCreate

    conteudo = await arquivo.read()
    nome = (arquivo.filename or "").lower()

    if nome.endswith(".xlsx") or nome.endswith(".xls"):
        linhas, cabecalho = ler_linhas_xlsx(conteudo)
    elif nome.endswith(".csv"):
        linhas, cabecalho = ler_linhas_csv(conteudo)
    else:
        raise HTTPException(status_code=400, detail="Formato não suportado. Envie .csv ou .xlsx")

    if not linhas:
        return {"criados": 0, "erros": 0, "total": 0, "mensagem": "Nenhuma linha de dados encontrada no arquivo."}

    col_data = resolver_coluna(cabecalho, "data", mapa_data)
    col_valor = resolver_coluna(cabecalho, "valor", mapa_valor)
    col_descricao = resolver_coluna(cabecalho, "descricao", mapa_descricao)
    col_tipo = resolver_coluna(cabecalho, "tipo", mapa_tipo)
    col_fallback_desc = resolver_coluna(cabecalho, "categoria_fallback", None)
    col_status = resolver_coluna(cabecalho, "status", None)

    if not col_data or not col_valor:
        return {
            "criados": 0, "erros": len(linhas), "total": len(linhas),
            "mensagem": (
                f"Não encontrei as colunas obrigatórias de data/valor automaticamente. "
                f"Cabeçalho detectado: {cabecalho}. Informe manualmente 'Coluna de Data' e 'Coluna de Valor' no modal."
            ),
        }

    criados = 0
    ignorados_planejamento = 0
    erros_lista = []

    for i, linha in enumerate(linhas, start=2):
        try:
            # Pula transações ainda não realizadas (planejamento/previsão) —
            # livro caixa/LCDPR só deve refletir movimentação de caixa real.
            # Só filtra se a coluna de status existir E o valor for reconhecido
            # como um status "não realizado"; valores desconhecidos não filtram
            # (evita descartar linha por engano num formato de arquivo diferente).
            if col_status:
                status_norm = normalizar_valor_status(linha.get(col_status))
                nao_realizado_hints = ("planejamento", "previsto", "pendente", "aguardando", "obrigatorio")
                if status_norm and status_norm not in STATUS_REALIZADOS and any(p in status_norm for p in nao_realizado_hints):
                    ignorados_planejamento += 1
                    continue

            data_lanc = parse_data(linha.get(col_data))
            valor = parse_valor(linha.get(col_valor))

            descricao = str(linha.get(col_descricao, "") or "").strip() if col_descricao else ""
            if (not descricao or descricao == "—") and col_fallback_desc:
                descricao = str(linha.get(col_fallback_desc, "") or "").strip()

            tipo_bruto = str(linha.get(col_tipo, "") or "").strip().lower() if col_tipo else ""
            if tipo_bruto in ("receita", "entrada", "credito"):
                tipo = "receita"
            elif tipo_bruto in ("despesa", "saida", "débito", "debito"):
                tipo = "despesa"
            elif valor is not None:
                # Sem coluna de tipo explicita/reconhecida: deriva do sinal do valor
                tipo = "receita" if valor > 0 else "despesa"
            else:
                tipo = "despesa"

            if not data_lanc:
                erros_lista.append(f"Linha {i}: data inválida ou ausente (valor recebido: {linha.get(col_data)!r})")
                continue
            if valor is None or valor == 0:
                erros_lista.append(f"Linha {i}: valor inválido ou ausente (valor recebido: {linha.get(col_valor)!r})")
                continue
            if not descricao:
                descricao = "Importação sem descrição"

            payload = LancamentoCreate(
                produtor_id=produtor_id,
                valor=abs(valor),
                data=data_lanc,
                data_lancamento=data_lanc,
                descricao=descricao,
                tipo=tipo,
                origem="importacao",
            )
            criar_lancamento(payload)
            criados += 1
        except Exception as e:
            erros_lista.append(f"Linha {i}: {str(e)}")

    return {
        "criados": criados,
        "erros": len(erros_lista),
        "ignorados_planejamento": ignorados_planejamento,
        "total": len(linhas),
        "mensagem": "; ".join(erros_lista[:10]) if erros_lista else (
            f"{ignorados_planejamento} linha(s) ignorada(s) por ainda não estarem realizadas (planejamento/previsão)."
            if ignorados_planejamento else None
        ),
    }
