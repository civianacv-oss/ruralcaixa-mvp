"""
RuralCaixa — routers/compravenda.py
Módulo Compra e Venda de Animais (atividade comercial não rural).

Adicione em app/main.py:
    from app.routers.compravenda import router as compravenda_router
    if compravenda_router: app.include_router(compravenda_router)
"""

from fastapi import APIRouter, HTTPException, Query
from pydantic import BaseModel, Field
from typing import List, Literal, Optional
from datetime import date, datetime
import psycopg2
import psycopg2.extras
import logging

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/compravenda", tags=["CompraVenda"])

DB_URL = "postgresql://postgres:tkyfcRsbrZuuHoThKgjuTiZWYVXOTdOX@gondola.proxy.rlwy.net:53900/railway"

def get_db():
    return psycopg2.connect(DB_URL, cursor_factory=psycopg2.extras.RealDictCursor)


# ─────────────────────────────────────────────────────────────
# SCHEMAS
# ─────────────────────────────────────────────────────────────

class ProdutoCreate(BaseModel):
    imovel_id: int
    nome: str
    descricao: Optional[str] = None
    unidade: str = "cab"          # cab, kg, arroba, saca, un
    especie: Optional[str] = None # bovino, suino, ovino, caprino, outro
    custo_medio: Optional[float] = None

class ProdutoUpdate(BaseModel):
    nome: Optional[str] = None
    descricao: Optional[str] = None
    unidade: Optional[str] = None
    especie: Optional[str] = None
    custo_medio: Optional[float] = None

class CompraCreate(BaseModel):
    imovel_id: int
    produto_id: int
    data_compra: date = Field(default_factory=date.today)
    quantidade: float
    valor_unitario: float
    fornecedor: Optional[str] = None
    nota_fiscal: Optional[str] = None
    observacoes: Optional[str] = None
    regime: Literal["pasto", "confinamento"] = "pasto"

class VendaCreate(BaseModel):
    imovel_id: int
    produto_id: int
    data_venda: date = Field(default_factory=date.today)
    quantidade: float
    valor_unitario: float
    comprador: Optional[str] = None
    nota_fiscal: Optional[str] = None
    observacoes: Optional[str] = None

class DespesaCreate(BaseModel):
    imovel_id: int
    descricao: str
    categoria: str = "operacional"   # operacional, logistica, administrativa, financeira
    data_lancamento: date = Field(default_factory=date.today)
    valor: float
    observacoes: Optional[str] = None


# ─────────────────────────────────────────────────────────────
# PRODUTOS / ESTOQUE
# ─────────────────────────────────────────────────────────────

@router.post("/produtos", status_code=201)
def criar_produto(dados: ProdutoCreate):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO cv_produtos (imovel_id, nome, descricao, unidade, especie, custo_medio)
            VALUES (%s, %s, %s, %s, %s, %s) RETURNING id
        """, (dados.imovel_id, dados.nome, dados.descricao, dados.unidade,
              dados.especie, dados.custo_medio))
        pid = cur.fetchone()["id"]
        conn.commit()
        return {"id": pid}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("/produtos")
def listar_produtos(imovel_id: int = Query(...)):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT p.*,
                   COALESCE(SUM(c.quantidade), 0) AS total_comprado,
                   COALESCE(SUM(v.quantidade), 0) AS total_vendido,
                   COALESCE(SUM(c.quantidade), 0) - COALESCE(SUM(v.quantidade), 0) AS estoque_atual,
                   CASE
                     WHEN COALESCE(SUM(c.quantidade), 0) > 0
                     THEN ROUND((COALESCE(SUM(c.valor_total), 0) / COALESCE(SUM(c.quantidade), 0))::numeric, 2)
                     ELSE p.custo_medio
                   END AS custo_medio_calc
            FROM cv_produtos p
            LEFT JOIN cv_compras c ON c.produto_id = p.id
            LEFT JOIN cv_vendas  v ON v.produto_id = p.id
            WHERE p.imovel_id = %s AND p.ativo = TRUE
            GROUP BY p.id
            ORDER BY p.nome
        """, (imovel_id,))
        return cur.fetchall()
    finally:
        conn.close()


@router.patch("/produtos/{produto_id}")
def atualizar_produto(produto_id: int, dados: ProdutoUpdate):
    conn = get_db()
    try:
        cur = conn.cursor()
        campos, valores = [], []
        for campo in ["nome", "descricao", "unidade", "especie", "custo_medio"]:
            val = getattr(dados, campo)
            if val is not None:
                campos.append(f"{campo} = %s")
                valores.append(val)
        if campos:
            valores.append(produto_id)
            cur.execute(f"UPDATE cv_produtos SET {', '.join(campos)} WHERE id = %s", valores)
            conn.commit()
        return {"ok": True}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────
# COMPRAS
# ─────────────────────────────────────────────────────────────

@router.post("/compras", status_code=201)
def registrar_compra(dados: CompraCreate):
    conn = get_db()
    try:
        cur = conn.cursor()
        valor_total = dados.quantidade * dados.valor_unitario
        cur.execute("""
            INSERT INTO cv_compras
                (imovel_id, produto_id, data_compra, quantidade, valor_unitario,
                 valor_total, fornecedor, nota_fiscal, observacoes, regime)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id
        """, (dados.imovel_id, dados.produto_id, dados.data_compra,
              dados.quantidade, dados.valor_unitario, valor_total,
              dados.fornecedor, dados.nota_fiscal, dados.observacoes,
              dados.regime))
        cid = cur.fetchone()["id"]
        conn.commit()
        return {"id": cid, "valor_total": valor_total}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("/compras")
def listar_compras(
    imovel_id: int = Query(...),
    produto_id: Optional[int] = Query(None),
    data_inicio: Optional[date] = Query(None),
    data_fim: Optional[date] = Query(None),
    limit: int = Query(100),
):
    conn = get_db()
    try:
        cur = conn.cursor()
        filtros = ["c.imovel_id = %s"]
        params: list = [imovel_id]
        if produto_id:
            filtros.append("c.produto_id = %s"); params.append(produto_id)
        if data_inicio:
            filtros.append("c.data_compra >= %s"); params.append(data_inicio)
        if data_fim:
            filtros.append("c.data_compra <= %s"); params.append(data_fim)
        params.append(limit)
        cur.execute(f"""
            SELECT c.*, p.nome AS produto_nome, p.unidade, p.especie
            FROM cv_compras c
            JOIN cv_produtos p ON p.id = c.produto_id
            WHERE {' AND '.join(filtros)}
            ORDER BY c.data_compra DESC
            LIMIT %s
        """, params)
        return cur.fetchall()
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────
# VENDAS
# ─────────────────────────────────────────────────────────────

PRAZO_FISCAL = {"confinamento": 52, "pasto": 138}  # Decreto 9.580/2018 — mesmo prazo do /alertas-fiscais

# Subcontas conhecidas de "Venda de {especie}" — mesmas UUIDs já usadas
# nos módulos zootécnicos (bovino.py etc). Se a espécie não estiver
# aqui, tenta achar dinamicamente na tabela subcontas pelo nome.
SUBCONTAS_VENDA_POR_ESPECIE = {
    "bovino": "93f28dea-0242-462c-a4ef-65eed3478815",
}


def _subconta_venda(cur, produto_id: int) -> str:
    """Resolve a subconta de receita a usar no lançamento de venda,
    a partir da espécie do produto. Lança HTTPException clara se não
    encontrar nenhuma opção — evita gravar num UUID inventado."""
    cur.execute("SELECT especie FROM cv_produtos WHERE id = %s", (produto_id,))
    row = cur.fetchone()
    especie = (row["especie"] if row else None) or ""
    especie = especie.lower()

    if especie in SUBCONTAS_VENDA_POR_ESPECIE:
        return SUBCONTAS_VENDA_POR_ESPECIE[especie]

    cur.execute("""
        SELECT id FROM subcontas
        WHERE LOWER(nome) LIKE %s AND LOWER(nome) LIKE '%%venda%%'
        LIMIT 1
    """, (f"%{especie}%",))
    achou = cur.fetchone()
    if achou:
        return achou["id"]

    raise HTTPException(
        500,
        f"Não encontrei uma subconta de 'Venda de {especie}' cadastrada. "
        f"Crie essa subconta no plano de contas antes de registrar a venda, "
        f"ou avise o desenvolvedor pra mapear a subconta certa pra essa espécie."
    )


def _registrar_venda_fifo(cur, imovel_id: int, produto_id: int, data_venda: date,
                           quantidade: float, valor_unitario: float,
                           comprador: str = None, nota_fiscal: str = None,
                           observacoes: str = None) -> dict:
    """
    Núcleo do módulo Compra e Venda: baixa o estoque por FIFO (Primeiro que
    Entra, Primeiro que Sai) e classifica cada porção da venda como RURAL
    ou NEGOCIACAO pela regra dos 52/138 dias (Decreto 9.580/2018).

    Não abre/fecha conexão nem faz commit — quem chama controla a
    transação (permite compor com outras operações, como o estorno de
    reclassificação, na mesma transação atômica).

    Reaproveitado por:
      - POST /vendas          (fluxo normal, produto já em cv_produtos)
      - POST /reclassificar-lancamento (migra um lançamento já feito
        direto no Livro Caixa — ex: bot classificou como investimento —
        pra dentro do módulo Compra e Venda, respeitando o mesmo prazo)
    """
    cur.execute("""
        SELECT c.id, c.data_compra, c.quantidade, c.valor_total, c.valor_unitario,
               c.regime,
               c.quantidade - COALESCE((
                   SELECT SUM(b.quantidade_baixada) FROM cv_vendas_baixas b
                   WHERE b.compra_id = c.id
               ), 0) AS saldo
        FROM cv_compras c
        WHERE c.produto_id = %s AND c.imovel_id = %s
        ORDER BY c.data_compra ASC, c.id ASC
    """, (produto_id, imovel_id))
    compras = [dict(r) for r in cur.fetchall() if float(r["saldo"]) > 0]

    estoque_disponivel = sum(float(c["saldo"]) for c in compras)
    if quantidade > estoque_disponivel:
        raise HTTPException(
            status_code=400,
            detail=f"Estoque insuficiente. Disponivel: {estoque_disponivel:.2f}"
        )

    valor_total = quantidade * valor_unitario
    restante = quantidade
    baixas = []

    for c in compras:
        if restante <= 0:
            break
        qtd_baixa = min(restante, float(c["saldo"]))
        dias = (data_venda - c["data_compra"]).days
        prazo_max = PRAZO_FISCAL.get(c["regime"], 138)
        classificacao = "RURAL" if dias > prazo_max else "NEGOCIACAO"

        valor_baixado = round(qtd_baixa * valor_unitario, 2)
        custo_baixado = round(qtd_baixa * float(c["valor_unitario"]), 2)

        baixas.append({
            "compra_id": c["id"],
            "quantidade_baixada": qtd_baixa,
            "dias_permanencia": dias,
            "prazo_max": prazo_max,
            "classificacao": classificacao,
            "valor_baixado": valor_baixado,
            "custo_baixado": custo_baixado,
        })
        restante -= qtd_baixa

    custo_total = sum(b["custo_baixado"] for b in baixas)
    lucro_bruto = valor_total - custo_total
    margem_pct = round((lucro_bruto / valor_total * 100), 2) if valor_total > 0 else 0

    valor_rural = sum(b["valor_baixado"] for b in baixas if b["classificacao"] == "RURAL")
    valor_negociacao = sum(b["valor_baixado"] for b in baixas if b["classificacao"] == "NEGOCIACAO")
    classificacoes_presentes = {b["classificacao"] for b in baixas}
    classificacao_venda = (
        classificacoes_presentes.pop() if len(classificacoes_presentes) == 1 else "MISTA"
    )

    # So a parte RURAL gera lancamento no Livro Caixa. A parte NEGOCIACAO
    # fica de fora e deve ser tratada separadamente na Declaracao Anual (DAA).
    #
    # IMPORTANTE: a tabela `lancamentos` real NAO tem conta_codigo/tipo/
    # descricao/imovel_id/data_lancamento (isso so existe no banco.sql de
    # referencia, desatualizado). O schema real e:
    #   propriedade_id, subconta_id (uuid), valor, data, origem,
    #   origem_modulo, origem_tipo, origem_id, origem_descricao.
    # A "conta" e definida pela subconta_id (tabela subcontas), nao por
    # um codigo textual.
    lancamento_id = None
    if valor_rural > 0:
        subconta_id = _subconta_venda(cur, produto_id)
        cur.execute("SELECT produtor_id FROM imoveis_rurais WHERE id = %s", (imovel_id,))
        row_produtor = cur.fetchone()
        produtor_id_lanc = row_produtor["produtor_id"] if row_produtor else None
        if not produtor_id_lanc:
            raise HTTPException(
                500, f"Não encontrei produtor_id para imovel_id={imovel_id} — "
                     f"não dá pra gravar o lançamento (coluna obrigatória)."
            )
        cur.execute("""
            INSERT INTO lancamentos
                (produtor_id, propriedade_id, subconta_id, valor, data, origem,
                 origem_modulo, origem_tipo, origem_descricao)
            VALUES (%s, %s, %s, %s, %s, 'compravenda', 'compravenda', 'venda', %s)
            RETURNING id
        """, (
            produtor_id_lanc,
            imovel_id,
            subconta_id,
            round(valor_rural, 2),
            data_venda,
            f"Venda comercial (parte rural) - {comprador or 'comprador nao informado'} - "
            f"classificado automaticamente pela regra dos 52/138 dias (Decreto 9.580/2018)",
        ))
        lancamento_id = cur.fetchone()["id"]

    cur.execute("""
        INSERT INTO cv_vendas
            (imovel_id, produto_id, data_venda, quantidade, valor_unitario,
             valor_total, custo_total, lucro_bruto, margem_pct,
             comprador, nota_fiscal, observacoes,
             classificacao, valor_rural, valor_negociacao, lancamento_id)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        RETURNING id
    """, (imovel_id, produto_id, data_venda,
          quantidade, valor_unitario, valor_total,
          custo_total, lucro_bruto, margem_pct,
          comprador, nota_fiscal, observacoes,
          classificacao_venda, round(valor_rural, 2), round(valor_negociacao, 2),
          lancamento_id))
    vid = cur.fetchone()["id"]

    if lancamento_id:
        cur.execute("UPDATE lancamentos SET origem_id = %s WHERE id = %s", (vid, lancamento_id))

    for b in baixas:
        cur.execute("""
            INSERT INTO cv_vendas_baixas
                (venda_id, compra_id, quantidade_baixada, dias_permanencia,
                 prazo_max, classificacao, valor_baixado, custo_baixado)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (vid, b["compra_id"], b["quantidade_baixada"], b["dias_permanencia"],
              b["prazo_max"], b["classificacao"], b["valor_baixado"], b["custo_baixado"]))

    return {
        "id": vid,
        "valor_total": valor_total,
        "lucro_bruto": lucro_bruto,
        "margem_pct": margem_pct,
        "classificacao": classificacao_venda,
        "valor_rural": round(valor_rural, 2),
        "valor_negociacao": round(valor_negociacao, 2),
        "lancamento_id": lancamento_id,
        "aviso": (
            None if classificacao_venda == "RURAL" else
            "Parte ou toda a venda ficou dentro do prazo fiscal (regime comercial) - "
            "nao entrou no Livro Caixa Rural. Declare o valor_negociacao separadamente "
            "como ganho de capital / atividade comercial na Declaracao Anual."
        ),
    }


@router.post("/vendas", status_code=201)
def registrar_venda(dados: VendaCreate):
    """
    Registra a venda e baixa o estoque por FIFO (Primeiro que Entra,
    Primeiro que Sai), consumindo as compras mais antigas primeiro.

    Cada porção baixada é classificada individualmente:
      - RURAL:      dias em estoque > prazo fiscal do regime da compra
                     (52d confinamento / 138d pasto) -> entra no Livro Caixa
      - NEGOCIACAO: dias em estoque <= prazo fiscal -> NAO entra no Livro
                     Caixa Rural; deve ser declarado a parte como ganho de
                     capital / atividade comercial na Declaracao Anual.

    Se a venda consumir compras de mais de uma classificacao (parte ja
    "rural", parte ainda "negociacao"), a venda fica com classificacao
    = 'MISTA' e o valor e dividido proporcionalmente entre valor_rural
    e valor_negociacao. So o valor_rural gera lancamento no Livro Caixa.
    """
    conn = get_db()
    try:
        cur = conn.cursor()
        resultado = _registrar_venda_fifo(
            cur, dados.imovel_id, dados.produto_id, dados.data_venda,
            dados.quantidade, dados.valor_unitario,
            dados.comprador, dados.nota_fiscal, dados.observacoes,
        )
        conn.commit()
        return resultado
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


class ReclassificarLancamento(BaseModel):
    lancamento_id: str            # UUID do lançamento já criado direto no Livro Caixa (errado)
    imovel_id: int
    produto_id: Optional[int] = None
    especie: Optional[str] = None  # usado só se produto_id não vier, pra achar/criar em cv_produtos
    quantidade: float = Field(gt=0)   # nº de animais do lançamento original / agora vendidos
    regime: Literal["pasto", "confinamento"] = "pasto"
    data_venda: date
    valor_venda: float = Field(ge=0)  # valor TOTAL recebido na venda (não unitário)
    comprador: Optional[str] = None
    nota_fiscal: Optional[str] = None


@router.post("/reclassificar-lancamento", status_code=201)
def reclassificar_lancamento(dados: ReclassificarLancamento):
    """
    Corrige um lançamento que foi criado direto no Livro Caixa (ex: bot
    classificou "compra de bezerro" como Investimento) mas que na verdade
    era destinado à revenda e acabou sendo vendido dentro do prazo fiscal
    (52 dias confinamento / 138 dias pasto — Decreto 9.580/2018).

    Faz três coisas, na mesma transação:
      1. Estorna o lançamento original (não apaga — cria um lançamento
         reverso de mesmo valor/conta, pra manter rastro de auditoria).
      2. Migra a compra pro módulo Compra e Venda (cv_compras), com a
         DATA ORIGINAL da compra — é essa data que conta pro prazo fiscal.
      3. Registra a venda via FIFO/classificação automática — se ainda
         estiver dentro do prazo, fica como NEGOCIACAO (fora do Livro
         Caixa Rural); se já passou, entra como RURAL normalmente.
    """
    conn = get_db()
    try:
        cur = conn.cursor()

        cur.execute(
            "SELECT id, valor, data, propriedade_id, subconta_id FROM lancamentos WHERE id = %s",
            (dados.lancamento_id,),
        )
        original = cur.fetchone()
        if not original:
            raise HTTPException(404, "Lançamento não encontrado")
        if original["propriedade_id"] and int(original["propriedade_id"]) != dados.imovel_id:
            raise HTTPException(400, "Lançamento pertence a outro imóvel")

        valor_original = abs(float(original["valor"]))
        data_compra_original = original["data"]

        # 1. Resolve produto_id (reaproveita por espécie, ou cria um novo)
        produto_id = dados.produto_id
        if not produto_id:
            if not dados.especie:
                raise HTTPException(400, "Informe produto_id ou especie")
            cur.execute("""
                SELECT id FROM cv_produtos
                WHERE imovel_id = %s AND LOWER(especie) = LOWER(%s) AND ativo = TRUE
                LIMIT 1
            """, (dados.imovel_id, dados.especie))
            existente = cur.fetchone()
            if existente:
                produto_id = existente["id"]
            else:
                cur.execute("""
                    INSERT INTO cv_produtos (imovel_id, nome, unidade, especie)
                    VALUES (%s, %s, 'cab', %s) RETURNING id
                """, (dados.imovel_id, dados.especie.capitalize(), dados.especie))
                produto_id = cur.fetchone()["id"]

        # 2. Estorna o lançamento original — lançamento reverso (valor
        #    negativo), mesma subconta/produtor, pra cancelar na mesma
        #    categoria e manter rastro de auditoria. A tabela real não
        #    tem conta_codigo/tipo/descricao (só existiam no banco.sql
        #    de referência, desatualizado) — usa subconta_id + valor.
        cur.execute("""
            INSERT INTO lancamentos
                (produtor_id, propriedade_id, subconta_id, valor, data, origem,
                 origem_modulo, origem_tipo, origem_id, origem_descricao)
            SELECT produtor_id, propriedade_id, subconta_id, -valor, %s,
                   'compravenda_estorno', 'compravenda', 'estorno', NULL,
                   'Lancamento original ' || id::text || ' reclassificado pela regra dos 52/138 dias'
            FROM lancamentos WHERE id = %s
            RETURNING id
        """, (date.today(), dados.lancamento_id))
        estorno_id = cur.fetchone()["id"]

        # 3. Migra a compra pro módulo Compra e Venda com a DATA ORIGINAL
        #    (é ela que determina se já passou do prazo fiscal ou não)
        valor_unitario_compra = valor_original / dados.quantidade
        cur.execute("""
            INSERT INTO cv_compras
                (imovel_id, produto_id, data_compra, quantidade, valor_unitario,
                 valor_total, regime, observacoes)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            dados.imovel_id, produto_id, data_compra_original, dados.quantidade,
            valor_unitario_compra, valor_original, dados.regime,
            f"Migrado do lançamento {dados.lancamento_id} (reclassificação fiscal)",
        ))
        compra_id = cur.fetchone()["id"]

        # 4. Registra a venda — o FIFO vai consumir exatamente essa compra
        #    recém-criada e classificar RURAL/NEGOCIACAO automaticamente
        resultado_venda = _registrar_venda_fifo(
            cur, dados.imovel_id, produto_id, dados.data_venda,
            dados.quantidade, dados.valor_venda / dados.quantidade,
            dados.comprador, dados.nota_fiscal,
            f"Reclassificado do lançamento original {dados.lancamento_id}",
        )

        conn.commit()
        return {
            "estorno_lancamento_id": estorno_id,
            "compra_id": compra_id,
            **resultado_venda,
        }
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("/vendas")
def listar_vendas(
    imovel_id: int = Query(...),
    produto_id: Optional[int] = Query(None),
    data_inicio: Optional[date] = Query(None),
    data_fim: Optional[date] = Query(None),
    limit: int = Query(100),
):
    conn = get_db()
    try:
        cur = conn.cursor()
        filtros = ["v.imovel_id = %s"]
        params: list = [imovel_id]
        if produto_id:
            filtros.append("v.produto_id = %s"); params.append(produto_id)
        if data_inicio:
            filtros.append("v.data_venda >= %s"); params.append(data_inicio)
        if data_fim:
            filtros.append("v.data_venda <= %s"); params.append(data_fim)
        params.append(limit)
        cur.execute(f"""
            SELECT v.*, p.nome AS produto_nome, p.unidade, p.especie
            FROM cv_vendas v
            JOIN cv_produtos p ON p.id = v.produto_id
            WHERE {' AND '.join(filtros)}
            ORDER BY v.data_venda DESC
            LIMIT %s
        """, params)
        return cur.fetchall()
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────
# DESPESAS OPERACIONAIS
# ─────────────────────────────────────────────────────────────

@router.post("/despesas", status_code=201)
def registrar_despesa(dados: DespesaCreate):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO cv_despesas (imovel_id, descricao, categoria, data_lancamento, valor, observacoes)
            VALUES (%s, %s, %s, %s, %s, %s) RETURNING id
        """, (dados.imovel_id, dados.descricao, dados.categoria,
              dados.data_lancamento, dados.valor, dados.observacoes))
        did = cur.fetchone()["id"]
        conn.commit()
        return {"id": did}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("/despesas")
def listar_despesas(
    imovel_id: int = Query(...),
    data_inicio: Optional[date] = Query(None),
    data_fim: Optional[date] = Query(None),
    limit: int = Query(100),
):
    conn = get_db()
    try:
        cur = conn.cursor()
        filtros = ["imovel_id = %s"]
        params: list = [imovel_id]
        if data_inicio:
            filtros.append("data_lancamento >= %s"); params.append(data_inicio)
        if data_fim:
            filtros.append("data_lancamento <= %s"); params.append(data_fim)
        params.append(limit)
        cur.execute(f"""
            SELECT * FROM cv_despesas
            WHERE {' AND '.join(filtros)}
            ORDER BY data_lancamento DESC
            LIMIT %s
        """, params)
        return cur.fetchall()
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────
# DASHBOARD / DRE / FLUXO DE CAIXA
# ─────────────────────────────────────────────────────────────

@router.get("/dashboard/{imovel_id}")
def dashboard(imovel_id: int):
    conn = get_db()
    try:
        cur = conn.cursor()

        # Receita bruta de vendas
        cur.execute("""
            SELECT COALESCE(SUM(valor_total), 0) AS receita_bruta,
                   COALESCE(SUM(lucro_bruto), 0) AS lucro_bruto_total,
                   COALESCE(SUM(custo_total), 0)  AS cmv_total,
                   COUNT(*) AS total_vendas
            FROM cv_vendas WHERE imovel_id = %s
        """, (imovel_id,))
        vendas_res = dict(cur.fetchone())

        # Custo de compras
        cur.execute("""
            SELECT COALESCE(SUM(valor_total), 0) AS total_compras,
                   COUNT(*) AS total_compras_qtd
            FROM cv_compras WHERE imovel_id = %s
        """, (imovel_id,))
        compras_res = dict(cur.fetchone())

        # Despesas operacionais
        cur.execute("""
            SELECT COALESCE(SUM(valor), 0) AS total_despesas
            FROM cv_despesas WHERE imovel_id = %s
        """, (imovel_id,))
        desp_res = dict(cur.fetchone())

        # Estoque atual (valor)
        cur.execute("""
            SELECT COALESCE(SUM(
                (COALESCE(c.qtd, 0) - COALESCE(v.qtd, 0)) *
                CASE WHEN COALESCE(c.qtd, 0) > 0
                     THEN c.val / c.qtd ELSE 0 END
            ), 0) AS valor_estoque
            FROM cv_produtos p
            LEFT JOIN (
                SELECT produto_id, SUM(quantidade) AS qtd, SUM(valor_total) AS val
                FROM cv_compras WHERE imovel_id = %s GROUP BY produto_id
            ) c ON c.produto_id = p.id
            LEFT JOIN (
                SELECT produto_id, SUM(quantidade) AS qtd
                FROM cv_vendas WHERE imovel_id = %s GROUP BY produto_id
            ) v ON v.produto_id = p.id
            WHERE p.imovel_id = %s AND p.ativo = TRUE
        """, (imovel_id, imovel_id, imovel_id))
        estoque_res = dict(cur.fetchone())

        # Margem por produto (top 10)
        cur.execute("""
            SELECT p.nome, p.especie, p.unidade,
                   COALESCE(SUM(v.quantidade), 0) AS qtd_vendida,
                   COALESCE(SUM(v.valor_total), 0) AS receita,
                   COALESCE(SUM(v.custo_total), 0) AS custo,
                   COALESCE(SUM(v.lucro_bruto), 0) AS lucro,
                   CASE WHEN COALESCE(SUM(v.valor_total), 0) > 0
                        THEN ROUND((COALESCE(SUM(v.lucro_bruto), 0) /
                             COALESCE(SUM(v.valor_total), 0) * 100)::numeric, 2)
                        ELSE 0 END AS margem_pct
            FROM cv_produtos p
            LEFT JOIN cv_vendas v ON v.produto_id = p.id
            WHERE p.imovel_id = %s AND p.ativo = TRUE
            GROUP BY p.id, p.nome, p.especie, p.unidade
            ORDER BY lucro DESC
            LIMIT 10
        """, (imovel_id,))
        margem_produtos = [dict(r) for r in cur.fetchall()]

        receita_bruta = float(vendas_res["receita_bruta"])
        cmv = float(vendas_res["cmv_total"])
        lucro_bruto = float(vendas_res["lucro_bruto_total"])
        despesas_op = float(desp_res["total_despesas"])
        lucro_liquido = lucro_bruto - despesas_op
        margem_bruta_pct = round(lucro_bruto / receita_bruta * 100, 2) if receita_bruta > 0 else 0
        margem_liquida_pct = round(lucro_liquido / receita_bruta * 100, 2) if receita_bruta > 0 else 0

        return {
            "resumo": {
                "receita_bruta": receita_bruta,
                "cmv": cmv,
                "lucro_bruto": lucro_bruto,
                "margem_bruta_pct": margem_bruta_pct,
                "despesas_operacionais": despesas_op,
                "lucro_liquido": lucro_liquido,
                "margem_liquida_pct": margem_liquida_pct,
                "total_investido_compras": float(compras_res["total_compras"]),
                "valor_estoque_atual": float(estoque_res["valor_estoque"]),
                "total_vendas": int(vendas_res["total_vendas"]),
                "total_compras": int(compras_res["total_compras_qtd"]),
            },
            "margem_por_produto": margem_produtos,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("/fluxo-caixa/{imovel_id}")
def fluxo_caixa(
    imovel_id: int,
    meses: int = Query(12),
):
    conn = get_db()
    try:
        cur = conn.cursor()

        # Entradas (vendas) por mês
        cur.execute("""
            SELECT TO_CHAR(data_venda, 'YYYY-MM') AS mes,
                   SUM(valor_total) AS entradas,
                   SUM(lucro_bruto) AS lucro_bruto
            FROM cv_vendas
            WHERE imovel_id = %s
              AND data_venda >= CURRENT_DATE - INTERVAL '%s months'
            GROUP BY mes ORDER BY mes
        """, (imovel_id, meses))
        entradas_rows = {r["mes"]: dict(r) for r in cur.fetchall()}

        # Saídas (compras) por mês
        cur.execute("""
            SELECT TO_CHAR(data_compra, 'YYYY-MM') AS mes,
                   SUM(valor_total) AS saidas_compras
            FROM cv_compras
            WHERE imovel_id = %s
              AND data_compra >= CURRENT_DATE - INTERVAL '%s months'
            GROUP BY mes ORDER BY mes
        """, (imovel_id, meses))
        compras_rows = {r["mes"]: dict(r) for r in cur.fetchall()}

        # Despesas por mês
        cur.execute("""
            SELECT TO_CHAR(data_lancamento, 'YYYY-MM') AS mes,
                   SUM(valor) AS saidas_despesas
            FROM cv_despesas
            WHERE imovel_id = %s
              AND data_lancamento >= CURRENT_DATE - INTERVAL '%s months'
            GROUP BY mes ORDER BY mes
        """, (imovel_id, meses))
        desp_rows = {r["mes"]: dict(r) for r in cur.fetchall()}

        # Consolidar meses
        todos_meses = sorted(set(list(entradas_rows.keys()) + list(compras_rows.keys()) + list(desp_rows.keys())))
        fluxo = []
        saldo_acumulado = 0.0
        for mes in todos_meses:
            entradas = float(entradas_rows.get(mes, {}).get("entradas") or 0)
            saidas_compras = float(compras_rows.get(mes, {}).get("saidas_compras") or 0)
            saidas_despesas = float(desp_rows.get(mes, {}).get("saidas_despesas") or 0)
            saidas = saidas_compras + saidas_despesas
            saldo_mes = entradas - saidas
            saldo_acumulado += saldo_mes
            fluxo.append({
                "mes": mes,
                "entradas": round(entradas, 2),
                "saidas_compras": round(saidas_compras, 2),
                "saidas_despesas": round(saidas_despesas, 2),
                "saidas_total": round(saidas, 2),
                "saldo_mes": round(saldo_mes, 2),
                "saldo_acumulado": round(saldo_acumulado, 2),
            })

        return fluxo
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("/dre/{imovel_id}")
def dre(
    imovel_id: int,
    ano: Optional[int] = Query(None),
):
    conn = get_db()
    try:
        cur = conn.cursor()
        ano_filtro = ano or datetime.now().year

        cur.execute("""
            SELECT
                COALESCE(SUM(valor_total), 0)  AS receita_bruta,
                COALESCE(SUM(custo_total), 0)  AS cmv,
                COALESCE(SUM(lucro_bruto), 0)  AS lucro_bruto
            FROM cv_vendas
            WHERE imovel_id = %s
              AND EXTRACT(YEAR FROM data_venda) = %s
        """, (imovel_id, ano_filtro))
        v = dict(cur.fetchone())

        cur.execute("""
            SELECT categoria, COALESCE(SUM(valor), 0) AS total
            FROM cv_despesas
            WHERE imovel_id = %s AND EXTRACT(YEAR FROM data_lancamento) = %s
            GROUP BY categoria
        """, (imovel_id, ano_filtro))
        despesas_cat = {r["categoria"]: float(r["total"]) for r in cur.fetchall()}
        total_despesas = sum(despesas_cat.values())

        receita_bruta = float(v["receita_bruta"])
        cmv = float(v["cmv"])
        lucro_bruto = float(v["lucro_bruto"])
        lucro_operacional = lucro_bruto - total_despesas

        return {
            "ano": ano_filtro,
            "receita_bruta": round(receita_bruta, 2),
            "cmv": round(cmv, 2),
            "lucro_bruto": round(lucro_bruto, 2),
            "margem_bruta_pct": round(lucro_bruto / receita_bruta * 100, 2) if receita_bruta > 0 else 0,
            "despesas": despesas_cat,
            "total_despesas": round(total_despesas, 2),
            "lucro_operacional": round(lucro_operacional, 2),
            "margem_operacional_pct": round(lucro_operacional / receita_bruta * 100, 2) if receita_bruta > 0 else 0,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────
# RELATÓRIO FISCAL — RURAL vs NEGOCIAÇÃO (regra dos 52/138 dias)
# ─────────────────────────────────────────────────────────────

@router.get("/relatorio-fiscal/{imovel_id}")
def relatorio_fiscal(
    imovel_id: int,
    data_inicio: Optional[date] = Query(None),
    data_fim: Optional[date] = Query(None),
):
    """
    Consolida as vendas por classificação fiscal (RURAL / NEGOCIACAO),
    já com a baixa por FIFO aplicada em /vendas. Serve de base para a
    tela "Compra e Venda" e para o contador separar o que vai no Livro
    Caixa Rural do que precisa ser declarado na DAA como negociação.
    """
    conn = get_db()
    try:
        cur = conn.cursor()
        filtros = ["v.imovel_id = %s"]
        params: list = [imovel_id]
        if data_inicio:
            filtros.append("v.data_venda >= %s"); params.append(data_inicio)
        if data_fim:
            filtros.append("v.data_venda <= %s"); params.append(data_fim)

        cur.execute(f"""
            SELECT b.classificacao,
                   COUNT(DISTINCT v.id) AS qtd_vendas,
                   SUM(b.valor_baixado) AS valor,
                   SUM(b.custo_baixado) AS custo,
                   SUM(b.valor_baixado - b.custo_baixado) AS resultado
            FROM cv_vendas_baixas b
            JOIN cv_vendas v ON v.id = b.venda_id
            WHERE {' AND '.join(filtros)}
            GROUP BY b.classificacao
        """, params)
        por_classificacao = {r["classificacao"]: dict(r) for r in cur.fetchall()}

        cur.execute(f"""
            SELECT v.id, p.nome AS produto_nome, p.especie, v.data_venda,
                   v.quantidade, v.classificacao, v.valor_rural, v.valor_negociacao,
                   v.lancamento_id
            FROM cv_vendas v
            JOIN cv_produtos p ON p.id = v.produto_id
            WHERE {' AND '.join(filtros)}
            ORDER BY v.data_venda DESC
        """, params)
        vendas = [dict(r) for r in cur.fetchall()]

        return {
            "resumo": {
                "rural": por_classificacao.get("RURAL", {"qtd_vendas": 0, "valor": 0, "custo": 0, "resultado": 0}),
                "negociacao": por_classificacao.get("NEGOCIACAO", {"qtd_vendas": 0, "valor": 0, "custo": 0, "resultado": 0}),
            },
            "vendas": vendas,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("/relatorio-ganho-capital/{imovel_id}")
def relatorio_ganho_capital(
    imovel_id: int,
    ano: Optional[int] = Query(None, description="Filtra pelo ano da venda (ano-calendário do GCAP)"),
):
    """
    Gera uma planilha (.xlsx) com as operações classificadas como
    NEGOCIACAO — as que ficaram FORA do Livro Caixa Rural por estarem
    dentro do prazo fiscal (52/138 dias) — com os campos necessários
    pra declarar Ganho de Capital.

    IMPORTANTE — limitação real da Receita Federal: o programa GCAP não
    aceita importação em lote. Cada operação precisa ser digitada
    manualmente lá (Identificação → Direitos/Bens Móveis → Novo). Esta
    planilha existe pra deixar essa digitação rápida (copiar e colar) ou
    pra entregar pronta ao contador — ela não gera o arquivo .DEC que o
    GCAP exporta pra declaração anual.

    Isso não é orientação tributária — confirme com um contador se a
    venda de animais de compra-e-venda se enquadra como "bem móvel" no
    GCAP no seu caso específico antes de declarar.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse

    conn = get_db()
    try:
        cur = conn.cursor()
        filtros = ["v.imovel_id = %s", "b.classificacao = 'NEGOCIACAO'"]
        params: list = [imovel_id]
        if ano:
            filtros.append("EXTRACT(YEAR FROM v.data_venda) = %s")
            params.append(ano)

        cur.execute(f"""
            SELECT p.especie, p.nome AS produto_nome,
                   c.data_compra, v.data_venda,
                   b.quantidade_baixada, b.dias_permanencia,
                   b.custo_baixado, b.valor_baixado,
                   (b.valor_baixado - b.custo_baixado) AS ganho_capital,
                   v.comprador
            FROM cv_vendas_baixas b
            JOIN cv_vendas v ON v.id = b.venda_id
            JOIN cv_compras c ON c.id = b.compra_id
            JOIN cv_produtos p ON p.id = v.produto_id
            WHERE {' AND '.join(filtros)}
            ORDER BY v.data_venda ASC
        """, params)
        linhas = cur.fetchall()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ganho de Capital"

    ws.append([f"Relatório de Ganho de Capital — Compra e Venda (fora do Livro Caixa Rural)"])
    ws.merge_cells("A1:J1")
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([
        "Operações classificadas como NEGOCIAÇÃO (dentro do prazo fiscal de 52/138 dias). "
        "Digite manualmente no GCAP em Identificação → Direitos/Bens Móveis → Novo."
    ])
    ws.merge_cells("A2:J2")
    ws["A2"].font = Font(italic=True, size=9, color="666666")
    ws.append([])

    headers = [
        "Espécie", "Produto/Lote", "Data de Aquisição", "Data de Alienação",
        "Quantidade", "Dias em Estoque", "Valor de Aquisição (R$)",
        "Valor de Alienação (R$)", "Ganho de Capital (R$)", "Comprador",
    ]
    ws.append(headers)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    total_aquisicao = total_alienacao = total_ganho = 0.0
    for l in linhas:
        custo = float(l["custo_baixado"])
        valor = float(l["valor_baixado"])
        ganho = float(l["ganho_capital"])
        ws.append([
            l["especie"], l["produto_nome"],
            l["data_compra"].strftime("%d/%m/%Y"),
            l["data_venda"].strftime("%d/%m/%Y"),
            float(l["quantidade_baixada"]), l["dias_permanencia"],
            round(custo, 2), round(valor, 2), round(ganho, 2),
            l["comprador"] or "",
        ])
        total_aquisicao += custo
        total_alienacao += valor
        total_ganho += ganho

    if not linhas:
        ws.append(["Nenhuma operação de negociação encontrada no período."])
        ws.merge_cells(f"A{ws.max_row}:J{ws.max_row}")
    else:
        ws.append([
            "TOTAL", "", "", "", "", "",
            round(total_aquisicao, 2), round(total_alienacao, 2), round(total_ganho, 2), "",
        ])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

    for col_cells in ws.columns:
        larguras = [len(str(c.value)) for c in col_cells if c.value is not None]
        if larguras:
            letra = col_cells[0].column_letter if hasattr(col_cells[0], "column_letter") else None
            if letra:
                ws.column_dimensions[letra].width = min(max(larguras) + 2, 32)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nome_arquivo = f"ganho_capital_imovel_{imovel_id}" + (f"_{ano}" if ano else "") + ".xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nome_arquivo}"},
    )


# ─────────────────────────────────────────────────────────────
# ALERTAS FISCAIS — animais próximos da transição para produção rural
# ─────────────────────────────────────────────────────────────

@router.get("/alertas-fiscais/{imovel_id}")
def alertas_fiscais(
    imovel_id: int,
    dias_aviso: int = Query(default=10, ge=1, le=52,
        description="Quantos dias antes do prazo emitir alerta (default: 10)"),
):
    """
    Retorna compras cujos animais estão prestes a ultrapassar o prazo fiscal
    do RIR/2018 (Decreto 9.580/2018):
      - Confinamento: 52 dias
      - Pasto / outros: 138 dias

    Inclui três níveis de urgência:
      - 'critico'  : prazo já atingido ou ultrapassado (≥ prazo)
      - 'urgente'  : faltam ≤ dias_aviso dias para o prazo
      - 'atencao'  : entre dias_aviso+1 e 75% do prazo (zona amarela)
    """
    PRAZO = {"confinamento": 52, "pasto": 138}
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT
                c.id,
                c.produto_id,
                p.nome          AS produto_nome,
                p.especie,
                p.unidade,
                c.data_compra,
                c.quantidade,
                c.valor_total,
                c.regime,
                c.fornecedor,
                (CURRENT_DATE - c.data_compra)::int AS dias_em_estoque
            FROM cv_compras c
            JOIN cv_produtos p ON p.id = c.produto_id
            WHERE c.imovel_id = %s
            ORDER BY c.data_compra ASC
        """, (imovel_id,))
        rows = cur.fetchall()

        alertas = []
        for r in rows:
            r = dict(r)
            prazo_max = PRAZO.get(r["regime"], 138)
            dias = r["dias_em_estoque"]
            dias_restantes = prazo_max - dias
            pct_prazo = round(dias / prazo_max * 100, 1)

            if dias >= prazo_max:
                nivel = "critico"
            elif dias_restantes <= dias_aviso:
                nivel = "urgente"
            elif dias >= prazo_max * 0.75:
                nivel = "atencao"
            else:
                continue  # dentro do prazo seguro — não inclui no alerta

            alertas.append({
                "id":             r["id"],
                "produto_id":     r["produto_id"],
                "produto_nome":   r["produto_nome"],
                "especie":        r["especie"],
                "unidade":        r["unidade"],
                "data_compra":    r["data_compra"].isoformat() if hasattr(r["data_compra"], "isoformat") else str(r["data_compra"]),
                "quantidade":     float(r["quantidade"]),
                "valor_total":    float(r["valor_total"]),
                "regime":         r["regime"],
                "fornecedor":     r["fornecedor"],
                "dias_em_estoque": dias,
                "prazo_max":      prazo_max,
                "dias_restantes": dias_restantes,
                "pct_prazo":      pct_prazo,
                "nivel":          nivel,  # 'critico' | 'urgente' | 'atencao'
                "mensagem": (
                    f"RECLASSIFICADO — {dias} dias em estoque (prazo: {prazo_max} dias)"
                    if nivel == "critico" else
                    f"Faltam {dias_restantes} dia(s) para reclassificação como Atividade Rural"
                ),
            })

        # Ordenar: crítico primeiro, depois urgente, depois atenção
        ordem = {"critico": 0, "urgente": 1, "atencao": 2}
        alertas.sort(key=lambda x: (ordem[x["nivel"]], x["dias_restantes"]))

        return {
            "total_alertas": len(alertas),
            "criticos":  sum(1 for a in alertas if a["nivel"] == "critico"),
            "urgentes":  sum(1 for a in alertas if a["nivel"] == "urgente"),
            "atencao":   sum(1 for a in alertas if a["nivel"] == "atencao"),
            "dias_aviso": dias_aviso,
            "alertas":   alertas,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()
