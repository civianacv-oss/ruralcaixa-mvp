"""
RuralCaixa — routers/bovino.py (psycopg2 síncrono)
Compatível com o padrão do ovino.py existente.
"""
from fastapi import APIRouter, HTTPException, Query, UploadFile, File, Form
from pydantic import BaseModel
from typing import Optional, List
from datetime import date, timedelta
import psycopg2
import psycopg2.errors
import psycopg2.extras
import os
import json
import logging

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/bovino", tags=["Bovino"])

DB_URL = os.getenv("DATABASE_URL", "postgresql://postgres:tkyfcRsbrZuuHoThKgjuTiZWYVXOTdOX@gondola.proxy.rlwy.net:53900/railway")

def get_db():
    return psycopg2.connect(DB_URL, cursor_factory=psycopg2.extras.RealDictCursor)

# ── SCHEMAS ──────────────────────────────────────────────────
class AnimalIn(BaseModel):
    imovel_id: int
    brinco: str
    nome: Optional[str] = None
    raca_id: Optional[int] = None
    raca_nome: Optional[str] = None  # usado na importação: resolve raca_id por nome/código se possível
    sexo: str
    aptidao_manejo: str
    categoria: str
    data_nascimento: Optional[date] = None
    peso_nascimento: Optional[float] = None
    mae_id: Optional[int] = None
    pai_id: Optional[int] = None
    lote_id: Optional[int] = None
    data_entrada: Optional[date] = None
    origem: str = "nascimento"
    valor_aquisicao: Optional[float] = None
    observacoes: Optional[str] = None
    # Genealogia (importação) — usados quando o pai/mãe NÃO está cadastrado
    # neste rebanho. Se mae_id/pai_id vierem preenchidos, têm prioridade.
    nome_pai: Optional[str] = None
    nome_mae: Optional[str] = None
    registro_pai_externo: Optional[str] = None
    registro_mae_externo: Optional[str] = None
    composicao_racial: Optional[str] = None

@router.get("/desempenho")
def desempenho_rebanho(imovel_id: int, dias: int = Query(30, ge=7, le=180)):
    conn = get_db()
    try:
        cur = conn.cursor()
        data_inicio = date.today() - timedelta(days=dias)

        cur.execute(
            """
            SELECT a.id, a.brinco, a.nome, a.aptidao_manejo, a.lote_id, l.nome AS lote_nome
            FROM bovino_animais a
            LEFT JOIN bovino_lotes l ON l.id = a.lote_id
            WHERE a.imovel_id = %s AND a.status = \'ativo\'
            """,
            (imovel_id,),
        )
        animais = cur.fetchall()

        resultados = []
        for a in animais:
            tipo = "leite" if a["aptidao_manejo"] == "leite" else "corte"
            if tipo == "leite":
                cur.execute(
                    """
                    SELECT COUNT(*) AS n, COALESCE(SUM(volume_l), 0) AS total
                    FROM bovino_ordenha
                    WHERE animal_id = %s AND data >= %s
                    """,
                    (a["id"], data_inicio),
                )
                row_leite = cur.fetchone()
                producao = float(row_leite["total"]) if row_leite["n"] > 0 else None
            else:
                cur.execute(
                    """
                    SELECT peso_kg FROM bovino_pesagens
                    WHERE animal_id = %s AND data >= %s
                    ORDER BY data ASC
                    """,
                    (a["id"], data_inicio),
                )
                pesagens = cur.fetchall()
                producao = (
                    float(pesagens[-1]["peso_kg"]) - float(pesagens[0]["peso_kg"])
                    if len(pesagens) >= 2
                    else None
                )

            metrica_dia = round(producao / dias, 3) if producao is not None else None
            resultados.append(
                {
                    "animal_id": a["id"],
                    "brinco": a["brinco"],
                    "nome": a["nome"],
                    "tipo": tipo,
                    "lote_id": a["lote_id"],
                    "lote_nome": a["lote_nome"],
                    "producao_periodo": round(producao, 2) if producao is not None else None,
                    "metrica_dia": metrica_dia,
                    "score": None,
                }
            )

        for tipo_grp in ("leite", "corte"):
            grupo = [r for r in resultados if r["tipo"] == tipo_grp and r["metrica_dia"] is not None]
            grupo.sort(key=lambda r: r["metrica_dia"])
            n = len(grupo)
            for i, r in enumerate(grupo):
                r["score"] = round((i / (n - 1)) * 100) if n > 1 else 50

        return resultados
    finally:
        conn.close()


class PesagemIn(BaseModel):
    animal_id: int
    data: date
    peso_kg: float
    motivo: str = "rotina"
    observacoes: Optional[str] = None

class ProducaoLeiteIn(BaseModel):
    imovel_id: int
    animal_id: Optional[int] = None
    lote_id: Optional[int] = None
    data: date
    turno: str = "total"
    volume_l: float
    gordura_pct: Optional[float] = None
    proteina_pct: Optional[float] = None
    destinacao: str = "venda"
    preco_litro: Optional[float] = None

class AbateIn(BaseModel):
    animal_id: int
    data: date
    tipo: str
    peso_vivo_kg: Optional[float] = None
    peso_carcaca_kg: Optional[float] = None
    preco_arroba: Optional[float] = None
    valor_total: Optional[float] = None
    comprador: Optional[str] = None
    nota_fiscal: Optional[str] = None
    observacoes: Optional[str] = None

class SanitarioIn(BaseModel):
    animal_id: Optional[int] = None
    lote_id: Optional[int] = None
    tipo: str
    produto: str
    dose_ml: Optional[float] = None
    via_aplicacao: Optional[str] = None
    data_aplicacao: date
    data_reforco: Optional[date] = None
    responsavel: Optional[str] = None
    custo_total: Optional[float] = None
    observacoes: Optional[str] = None

class ReproducaoIn(BaseModel):
    femea_id: int
    touro_id: Optional[int] = None
    metodo: str
    data_cobertura: date
    observacoes: Optional[str] = None

class LoteIn(BaseModel):
    imovel_id: int
    nome: str
    aptidao: str
    piquete: Optional[str] = None
    capacidade: Optional[int] = None

# ── RAÇAS ────────────────────────────────────────────────────
@router.get("/racas")
def listar_racas(aptidao: Optional[str] = None):
    conn = get_db()
    try:
        cur = conn.cursor()
        if aptidao:
            cur.execute("SELECT * FROM bovino_racas WHERE ativo = TRUE AND aptidao = %s ORDER BY nome", (aptidao,))
        else:
            cur.execute("SELECT * FROM bovino_racas WHERE ativo = TRUE ORDER BY aptidao, nome")
        return list(cur.fetchall())
    finally:
        conn.close()

# ── LOTES ────────────────────────────────────────────────────
@router.get("/lotes/{imovel_id}")
def listar_lotes(imovel_id: int):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT l.*, COUNT(a.id) AS qtd_animais
            FROM bovino_lotes l
            LEFT JOIN bovino_animais a ON a.lote_id = l.id AND a.status = 'ativo'
            WHERE l.imovel_id = %s AND l.ativo = TRUE
            GROUP BY l.id ORDER BY l.nome
        """, (imovel_id,))
        return list(cur.fetchall())
    finally:
        conn.close()

@router.post("/lotes")
def criar_lote(data: LoteIn):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO bovino_lotes (imovel_id, nome, aptidao, piquete, capacidade)
            VALUES (%s,%s,%s,%s,%s) RETURNING *
        """, (data.imovel_id, data.nome, data.aptidao, data.piquete, data.capacidade))
        conn.commit()
        return dict(cur.fetchone())
    finally:
        conn.close()

# ── ANIMAIS ──────────────────────────────────────────────────
@router.get("/animais/{imovel_id}")
def listar_animais(
    imovel_id: int,
    aptidao: Optional[str] = None,
    status: str = "ativo",
    categoria: Optional[str] = None,
    lote_id: Optional[int] = None
):
    conn = get_db()
    try:
        cur = conn.cursor()
        conds = ["a.imovel_id = %s", "a.status = %s"]
        params = [imovel_id, status]
        if aptidao:
            conds.append("a.aptidao_manejo = %s"); params.append(aptidao)
        if categoria:
            conds.append("a.categoria = %s"); params.append(categoria)
        if lote_id:
            conds.append("a.lote_id = %s"); params.append(lote_id)
        where = " AND ".join(conds)
        cur.execute(f"""
            SELECT a.*, r.nome AS raca_nome, l.nome AS lote_nome,
                   p.peso_kg AS ultimo_peso, p.data AS data_ultimo_peso
            FROM bovino_animais a
            LEFT JOIN bovino_racas r ON r.id = a.raca_id
            LEFT JOIN bovino_lotes l ON l.id = a.lote_id
            LEFT JOIN LATERAL (
                SELECT peso_kg, data FROM bovino_pesagens
                WHERE animal_id = a.id ORDER BY data DESC LIMIT 1
            ) p ON TRUE
            WHERE {where} ORDER BY a.brinco
        """, params)
        return list(cur.fetchall())
    finally:
        conn.close()

@router.post("/animais")
def cadastrar_animal(data: AnimalIn):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT id FROM especie WHERE codigo = 'BOVINO'")
        especie_id = cur.fetchone()['id']

        raca_id = data.raca_id
        if raca_id is None and data.raca_nome:
            cur.execute(
                "SELECT id FROM bovino_racas WHERE LOWER(nome) = LOWER(%s) LIMIT 1",
                (data.raca_nome,)
            )
            row = cur.fetchone()
            raca_id = row["id"] if row else None
            # Não força criação de raça nova nem tenta match parcial (arriscado
            # dar palpite errado) — se não achar, composicao_racial guarda o
            # texto original mesmo assim, nada se perde.

        cur.execute("""
            INSERT INTO bovino_animais
            (imovel_id, especie_id, brinco, nome, raca_id, sexo, aptidao_manejo,
             categoria, data_nascimento, peso_nascimento, mae_id, pai_id, lote_id,
             data_entrada, origem, valor_aquisicao, observacoes,
             nome_pai, nome_mae, registro_pai_externo, registro_mae_externo, composicao_racial)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *
        """, (data.imovel_id, especie_id, data.brinco, data.nome, raca_id,
              data.sexo, data.aptidao_manejo, data.categoria, data.data_nascimento,
              data.peso_nascimento, data.mae_id, data.pai_id, data.lote_id,
              data.data_entrada or date.today(), data.origem,
              data.valor_aquisicao, data.observacoes,
              data.nome_pai, data.nome_mae, data.registro_pai_externo,
              data.registro_mae_externo, data.composicao_racial))
        conn.commit()
        return dict(cur.fetchone())
    finally:
        conn.close()


@router.post("/animais/relink-genealogia/{imovel_id}")
def relink_genealogia(imovel_id: int):
    """
    Segunda passada da importação de genealogia: tenta linkar pai_id/mae_id
    de verdade (FK) sempre que o registro_pai_externo/registro_mae_externo
    de um animal bater com o brinco de outro animal já cadastrado no mesmo
    rebanho (inclusive animais recém-importados no mesmo lote).
    Roda quantas vezes quiser — é idempotente (só atualiza o que ainda não
    está linkado).
    """
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            UPDATE bovino_animais AS filho
            SET pai_id = pai.id
            FROM bovino_animais AS pai
            WHERE filho.imovel_id = %s
              AND filho.pai_id IS NULL
              AND filho.registro_pai_externo IS NOT NULL
              AND pai.imovel_id = filho.imovel_id
              AND LOWER(pai.brinco) = LOWER(filho.registro_pai_externo)
            RETURNING filho.id
        """, (imovel_id,))
        pais_linkados = len(cur.fetchall())

        cur.execute("""
            UPDATE bovino_animais AS filho
            SET mae_id = mae.id
            FROM bovino_animais AS mae
            WHERE filho.imovel_id = %s
              AND filho.mae_id IS NULL
              AND filho.registro_mae_externo IS NOT NULL
              AND mae.imovel_id = filho.imovel_id
              AND LOWER(mae.brinco) = LOWER(filho.registro_mae_externo)
            RETURNING filho.id
        """, (imovel_id,))
        maes_linkadas = len(cur.fetchall())

        conn.commit()
        return {"pais_linkados": pais_linkados, "maes_linkadas": maes_linkadas}
    finally:
        conn.close()

class AnimalPatchIn(BaseModel):
    """Edição parcial — só os campos enviados são atualizados. Espelha o
    contrato real da mutation tRPC `updateAnimal` (server/routers/railway.ts),
    que manda PATCH com brinco/nome/raca/sexo/observacoes, todos opcionais."""
    brinco: Optional[str] = None
    nome: Optional[str] = None
    raca: Optional[str] = None
    sexo: Optional[str] = None
    observacoes: Optional[str] = None

@router.patch("/animais/{animal_id}")
def editar_animal(animal_id: int, data: AnimalPatchIn):
    """Edita parcialmente os dados cadastrais de um animal já existente
    (não altera status — isso é feito por PATCH /animais/{id}/status)."""
    conn = get_db()
    try:
        cur = conn.cursor()
        campos = []
        valores = []

        if data.brinco is not None:
            campos.append("brinco = %s"); valores.append(data.brinco)
        if data.nome is not None:
            campos.append("nome = %s"); valores.append(data.nome)
        if data.sexo is not None:
            campos.append("sexo = %s"); valores.append(data.sexo)
        if data.observacoes is not None:
            campos.append("observacoes = %s"); valores.append(data.observacoes)
        if data.raca is not None:
            cur.execute(
                "SELECT id FROM bovino_racas WHERE LOWER(nome) = LOWER(%s) LIMIT 1",
                (data.raca,)
            )
            row_raca = cur.fetchone()
            campos.append("raca_id = %s")
            valores.append(row_raca["id"] if row_raca else None)

        if not campos:
            raise HTTPException(400, "Nenhum campo informado.")

        campos.append("updated_at = NOW()")
        valores.append(animal_id)

        cur.execute(
            f"UPDATE bovino_animais SET {', '.join(campos)} WHERE id = %s RETURNING *",
            valores,
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Animal não encontrado")
        conn.commit()
        return dict(row)
    except HTTPException:
        raise
    except psycopg2.errors.UniqueViolation:
        conn.rollback()
        raise HTTPException(409, f"Já existe um animal com o brinco '{data.brinco}' neste imóvel")
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, str(e))
    finally:
        conn.close()

@router.patch("/animais/{animal_id}/status")
def atualizar_status(animal_id: int, status: str):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("UPDATE bovino_animais SET status=%s, updated_at=NOW() WHERE id=%s RETURNING id, status", (status, animal_id))
        conn.commit()
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Animal não encontrado")
        return dict(row)
    finally:
        conn.close()

# ── PESAGENS ─────────────────────────────────────────────────
@router.post("/pesagens")
def registrar_pesagem(data: PesagemIn):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO bovino_pesagens (animal_id, data, peso_kg, motivo, observacoes)
            VALUES (%s,%s,%s,%s,%s) RETURNING *
        """, (data.animal_id, data.data, data.peso_kg, data.motivo, data.observacoes))
        cur.execute("UPDATE bovino_animais SET updated_at=NOW() WHERE id=%s", (data.animal_id,))
        conn.commit()
        return dict(cur.fetchone()) if cur.rowcount else {}
    finally:
        conn.close()

@router.get("/pesagens/{animal_id}")
def historico_pesagens(animal_id: int):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT * FROM bovino_pesagens WHERE animal_id = %s ORDER BY data", (animal_id,))
        return list(cur.fetchall())
    finally:
        conn.close()

# ── LEITE ────────────────────────────────────────────────────
@router.post("/leite/producao")
def registrar_producao(data: ProducaoLeiteIn):
    if not data.animal_id and not data.lote_id:
        raise HTTPException(400, "Informe animal_id ou lote_id")
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO bovino_producao_leite
            (imovel_id, animal_id, lote_id, data, turno, volume_l,
             gordura_pct, proteina_pct, destinacao, preco_litro)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *
        """, (data.imovel_id, data.animal_id, data.lote_id, data.data,
              data.turno, data.volume_l, data.gordura_pct, data.proteina_pct,
              data.destinacao, data.preco_litro))
        conn.commit()
        return dict(cur.fetchone())
    finally:
        conn.close()

@router.get("/leite/producao/{imovel_id}")
def listar_producao(imovel_id: int, dias: int = Query(30, ge=1, le=365)):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT * FROM bovino_producao_leite
            WHERE imovel_id = %s AND data >= CURRENT_DATE - %s
            ORDER BY data DESC
        """, (imovel_id, dias))
        return list(cur.fetchall())
    finally:
        conn.close()

@router.get("/leite/resumo/{imovel_id}")
def resumo_leite(imovel_id: int, meses: int = 6):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT * FROM vw_producao_leite_mensal
            WHERE imovel_id = %s
              AND mes >= DATE_TRUNC('month', CURRENT_DATE - (%s * INTERVAL '1 month'))
            ORDER BY mes
        """, (imovel_id, meses))
        return list(cur.fetchall())
    finally:
        conn.close()

# ── ABATES ───────────────────────────────────────────────────
@router.post("/abates")
def registrar_abate(data: AbateIn):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO bovino_abates
            (animal_id, data, tipo, peso_vivo_kg, peso_carcaca_kg,
             preco_arroba, valor_total, comprador, nota_fiscal, observacoes)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *
        """, (data.animal_id, data.data, data.tipo, data.peso_vivo_kg,
              data.peso_carcaca_kg, data.preco_arroba, data.valor_total,
              data.comprador, data.nota_fiscal, data.observacoes))
        MAPA_STATUS_BAIXA = {
            "abate_proprio": "abatido",
            "abate_frigorif": "abatido",
            "venda": "vendido",
            "morte": "morto",
            "doacao": "descartado",
            "permuta": "descartado",
        }
        novo_status = MAPA_STATUS_BAIXA.get(data.tipo, "descartado")
        cur.execute("UPDATE bovino_animais SET status=%s, updated_at=NOW() WHERE id=%s", (novo_status, data.animal_id))
        conn.commit()
        return dict(cur.fetchone()) if cur.rowcount else {}
    finally:
        conn.close()

@router.get("/abates/{imovel_id}")
def listar_abates(imovel_id: int, dias: int = 90):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT ab.*, a.brinco, a.nome AS animal_nome, a.categoria
            FROM bovino_abates ab
            JOIN bovino_animais a ON a.id = ab.animal_id
            WHERE a.imovel_id = %s AND ab.data >= CURRENT_DATE - %s
            ORDER BY ab.data DESC
        """, (imovel_id, dias))
        return list(cur.fetchall())
    finally:
        conn.close()

# ── SANITÁRIO ────────────────────────────────────────────────
@router.post("/sanitario")
def registrar_sanitario(data: SanitarioIn):
    if not data.animal_id and not data.lote_id:
        raise HTTPException(400, "Informe animal_id ou lote_id")
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO bovino_sanitario
            (animal_id, lote_id, tipo, produto, dose_ml, via_aplicacao,
             data_aplicacao, data_reforco, responsavel, custo_total, observacoes)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *
        """, (data.animal_id, data.lote_id, data.tipo, data.produto, data.dose_ml,
              data.via_aplicacao, data.data_aplicacao, data.data_reforco,
              data.responsavel, data.custo_total, data.observacoes))
        conn.commit()
        return dict(cur.fetchone())
    finally:
        conn.close()

@router.get("/sanitario/{imovel_id}/proximos")
def proximos_reforcos(imovel_id: int, dias: int = 30):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT s.*, a.brinco, a.nome AS animal_nome, l.nome AS lote_nome
            FROM bovino_sanitario s
            LEFT JOIN bovino_animais a ON a.id = s.animal_id
            LEFT JOIN bovino_lotes l ON l.id = s.lote_id
            WHERE (a.imovel_id = %s OR l.imovel_id = %s)
              AND s.data_reforco BETWEEN CURRENT_DATE AND CURRENT_DATE + %s
            ORDER BY s.data_reforco
        """, (imovel_id, imovel_id, dias))
        return list(cur.fetchall())
    finally:
        conn.close()

# ── REPRODUÇÃO ───────────────────────────────────────────────
@router.post("/reproducao")
def registrar_cobertura(data: ReproducaoIn):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO bovino_reproducao (femea_id, touro_id, metodo, data_cobertura, observacoes)
            VALUES (%s,%s,%s,%s,%s) RETURNING *
        """, (data.femea_id, data.touro_id, data.metodo, data.data_cobertura, data.observacoes))
        conn.commit()
        return dict(cur.fetchone())
    finally:
        conn.close()

@router.get("/reproducao/{imovel_id}/prenhas")
def femeas_prenhas(imovel_id: int):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT r.*, a.brinco, a.nome AS femea_nome,
                   r.data_parto_prev,
                   (r.data_parto_prev - CURRENT_DATE) AS dias_para_parto
            FROM bovino_reproducao r
            JOIN bovino_animais a ON a.id = r.femea_id
            WHERE a.imovel_id = %s
              AND r.resultado = 'positivo'
              AND r.data_parto_real IS NULL
            ORDER BY r.data_parto_prev
        """, (imovel_id,))
        return list(cur.fetchall())
    finally:
        conn.close()

# ── DASHBOARD ────────────────────────────────────────────────
@router.get("/dashboard/{imovel_id}")
def dashboard(imovel_id: int):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT * FROM vw_rebanho_atual WHERE imovel_id = %s", (imovel_id,))
        rebanho = list(cur.fetchall())

        cur.execute("""
            SELECT
              COUNT(*) FILTER (WHERE aptidao_manejo='corte') AS total_corte,
              COUNT(*) FILTER (WHERE aptidao_manejo='leite') AS total_leite,
              COUNT(*) AS total_geral
            FROM bovino_animais WHERE imovel_id=%s AND status='ativo'
        """, (imovel_id,))
        totais = dict(cur.fetchone())

        cur.execute("""
            SELECT COALESCE(SUM(volume_l),0) AS volume_l, COALESCE(SUM(valor_total),0) AS receita
            FROM bovino_producao_leite
            WHERE imovel_id=%s AND data >= CURRENT_DATE - 30 AND destinacao='venda'
        """, (imovel_id,))
        leite_30d = dict(cur.fetchone())

        cur.execute("""
            SELECT COUNT(*) FROM bovino_sanitario s
            LEFT JOIN bovino_animais a ON a.id = s.animal_id
            LEFT JOIN bovino_lotes l ON l.id = s.lote_id
            WHERE (a.imovel_id=%s OR l.imovel_id=%s)
              AND s.data_reforco BETWEEN CURRENT_DATE AND CURRENT_DATE+30
        """, (imovel_id, imovel_id))
        reforcos = cur.fetchone()['count']

        cur.execute("""
            SELECT COUNT(*) FROM bovino_reproducao r
            JOIN bovino_animais a ON a.id=r.femea_id
            WHERE a.imovel_id=%s AND r.resultado='positivo' AND r.data_parto_real IS NULL
        """, (imovel_id,))
        prenhas = cur.fetchone()['count']

        return {
            "rebanho_por_categoria": rebanho,
            "totais": totais,
            "leite_30d": leite_30d,
            "alertas": {"reforcos_sanitarios": reforcos, "femeas_prenhas": prenhas}
        }
    finally:
        conn.close()

# ══════════════════════════════════════════════════════════════
# SCHEMAS — Gado Leiteiro
# ══════════════════════════════════════════════════════════════
class OrdenhaIn(BaseModel):
    imovel_id: int
    animal_id: Optional[int] = None
    lote_id: Optional[int] = None
    data: date
    turno: str = "total"
    volume_l: float
    gordura_pct: Optional[float] = None
    proteina_pct: Optional[float] = None
    ccs: Optional[int] = None
    ufc: Optional[int] = None
    destinacao: str = "venda"
    preco_litro: Optional[float] = None
    observacoes: Optional[str] = None

class IatfIn(BaseModel):
    imovel_id: int
    femea_id: int
    lote_id: Optional[int] = None
    protocolo: str
    data_inicio: date
    data_iatf: Optional[date] = None
    touro_id: Optional[int] = None
    semen_touro: Optional[str] = None
    tecnico: Optional[str] = None
    resultado: str = "aguardando"
    data_diagnostico: Optional[date] = None
    observacoes: Optional[str] = None

class DietaTransicaoIn(BaseModel):
    imovel_id: int
    animal_id: int
    fase: str
    data_inicio: date
    data_fim: Optional[date] = None
    dieta_descricao: str
    volumoso_kg_dia: Optional[float] = None
    concentrado_kg_dia: Optional[float] = None
    suplemento: Optional[str] = None
    responsavel: Optional[str] = None
    observacoes: Optional[str] = None

# ══════════════════════════════════════════════════════════════
# SCHEMAS — Gado de Corte
# ══════════════════════════════════════════════════════════════
class ConfinamentoIn(BaseModel):
    imovel_id: int
    lote_id: int
    data_entrada: date
    data_saida_prev: Optional[date] = None
    peso_entrada_kg: Optional[float] = None
    dieta: Optional[str] = None
    custo_diario_cab: Optional[float] = None
    objetivo: str = "terminacao"
    observacoes: Optional[str] = None

class ConfinamentoFechamentoIn(BaseModel):
    data_saida_real: date
    peso_saida_kg: Optional[float] = None
    status: str = "encerrado"

class ClassificacaoCarcacaIn(BaseModel):
    imovel_id: int
    animal_id: int
    abate_id: Optional[int] = None
    data: date
    frigorifico: Optional[str] = None
    maturidade: Optional[str] = None
    acabamento: Optional[str] = None
    conformacao: Optional[str] = None
    peso_carcaca_kg: Optional[float] = None
    rendimento_pct: Optional[float] = None
    preco_arroba: Optional[float] = None
    valor_total: Optional[float] = None
    nota_fiscal: Optional[str] = None
    observacoes: Optional[str] = None

class CustoProducaoIn(BaseModel):
    imovel_id: int
    lote_id: Optional[int] = None
    confinamento_id: Optional[int] = None
    periodo_inicio: date
    periodo_fim: Optional[date] = None
    categoria: str
    descricao: Optional[str] = None
    valor: float

# ══════════════════════════════════════════════════════════════
# ENDPOINTS — tipo_bovino do imóvel
# ══════════════════════════════════════════════════════════════
@router.get("/tipo/{imovel_id}")
def tipo_bovino(imovel_id: int):
    """Retorna o tipo de exploração bovina do imóvel: leite | corte | misto."""
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT COALESCE(tipo_bovino,'corte') AS tipo FROM imoveis_rurais WHERE id=%s", (imovel_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Imóvel não encontrado")
        return {"tipo_bovino": row["tipo"]}
    finally:
        conn.close()

@router.patch("/tipo/{imovel_id}")
def atualizar_tipo_bovino(imovel_id: int, tipo_bovino: str):
    """Atualiza o tipo de exploração bovina do imóvel."""
    if tipo_bovino not in ("leite", "corte", "misto"):
        raise HTTPException(400, "tipo_bovino deve ser leite, corte ou misto")
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(
            "UPDATE imoveis_rurais SET tipo_bovino=%s WHERE id=%s RETURNING id",
            (tipo_bovino, imovel_id)
        )
        if not cur.fetchone():
            raise HTTPException(404, "Imóvel não encontrado")
        conn.commit()
        return {"ok": True, "tipo_bovino": tipo_bovino}
    finally:
        conn.close()

# ══════════════════════════════════════════════════════════════
# ENDPOINTS — Gado Leiteiro: Ordenha
# ══════════════════════════════════════════════════════════════
@router.post("/leiteiro/ordenha")
def registrar_ordenha(data: OrdenhaIn):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO bovino_ordenha
              (imovel_id, animal_id, lote_id, data, turno, volume_l,
               gordura_pct, proteina_pct, ccs, ufc, destinacao, preco_litro, observacoes)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id, valor_total
        """, (
            data.imovel_id, data.animal_id, data.lote_id, data.data,
            data.turno, data.volume_l, data.gordura_pct, data.proteina_pct,
            data.ccs, data.ufc, data.destinacao, data.preco_litro, data.observacoes
        ))
        row = dict(cur.fetchone())
        conn.commit()
        return {"ok": True, **row}
    finally:
        conn.close()

@router.get("/leiteiro/ordenha/{imovel_id}")
def listar_ordenha(imovel_id: int, dias: int = Query(30, ge=1, le=365)):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT o.*, a.brinco, a.nome AS nome_animal, bl.nome AS nome_lote
            FROM bovino_ordenha o
            LEFT JOIN bovino_animais a ON a.id = o.animal_id
            LEFT JOIN bovino_lotes bl ON bl.id = o.lote_id
            WHERE o.imovel_id=%s AND o.data >= CURRENT_DATE - %s
            ORDER BY o.data DESC, o.turno
        """, (imovel_id, dias))
        return list(cur.fetchall())
    finally:
        conn.close()

from typing import List, Optional
from datetime import date
from pydantic import BaseModel


class OrdenhaImportItem(BaseModel):
    animal_id: int
    data: date
    volume_l: Optional[float] = None
    gordura_pct: Optional[float] = None
    proteina_pct: Optional[float] = None
    lactose_pct: Optional[float] = None
    es_pct: Optional[float] = None
    ccs: Optional[int] = None
    numero_ordenhas_dia: Optional[int] = None
    numero_controle_externo: Optional[int] = None


class OrdenhaImportIn(BaseModel):
    imovel_id: int
    itens: List[OrdenhaImportItem]


@router.post("/leiteiro/ordenha/importar")
def importar_ordenha(data: OrdenhaImportIn):
    conn = get_db()
    criados = 0
    duplicados = []
    erros = []
    try:
        cur = conn.cursor()
        for item in data.itens:
            cur.execute("SAVEPOINT sp_item")
            try:
                cur.execute(
                    """
                    INSERT INTO bovino_ordenha
                        (imovel_id, animal_id, data, volume_l, gordura_pct,
                         proteina_pct, lactose_pct, es_pct, ccs,
                         numero_ordenhas_dia, numero_controle_externo,
                         turno, destinacao, fonte)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'total','venda','gisleite')
                    """,
                    (
                        data.imovel_id, item.animal_id, item.data, item.volume_l,
                        item.gordura_pct, item.proteina_pct, item.lactose_pct,
                        item.es_pct, item.ccs, item.numero_ordenhas_dia,
                        item.numero_controle_externo,
                    ),
                )
                cur.execute("RELEASE SAVEPOINT sp_item")
                criados += 1
            except psycopg2.errors.UniqueViolation:
                cur.execute("ROLLBACK TO SAVEPOINT sp_item")
                duplicados.append({
                    "animal_id": item.animal_id,
                    "data": str(item.data),
                    "motivo": "Ja existe um registro de ordenha para este animal nesta data.",
                })
                continue
            except Exception as e:
                cur.execute("ROLLBACK TO SAVEPOINT sp_item")
                erros.append({"animal_id": item.animal_id, "data": str(item.data), "erro": str(e)})
                continue
        conn.commit()
        return {
            "ok": True,
            "criados": criados,
            "duplicados": duplicados,
            "erros": erros,
            "total": len(data.itens),
        }
    finally:
        conn.close()


class LactacaoImportItem(BaseModel):
    animal_id: int
    ordem_parto: Optional[int] = None
    data_parto: date
    duracao_lactacao_dias: Optional[int] = None
    producao_total_litros: Optional[float] = None
    producao_305d_litros: Optional[float] = None
    producao_acumulada_gordura: Optional[float] = None
    producao_acumulada_proteina: Optional[float] = None
    escore_corporal: Optional[float] = None
    raca_registro: Optional[str] = None
    ccs_media: Optional[int] = None
    data_encerramento: Optional[date] = None
    causa_encerramento: Optional[str] = None


class LactacaoImportIn(BaseModel):
    imovel_id: int
    itens: List[LactacaoImportItem]


@router.post("/leiteiro/lactacoes/importar")
def importar_lactacoes(data: LactacaoImportIn):
    conn = get_db()
    criados = 0
    duplicados = []
    erros = []
    try:
        cur = conn.cursor()
        for item in data.itens:
            cur.execute("SAVEPOINT sp_item")
            try:
                cur.execute(
                    """
                    INSERT INTO bovino_lactacoes
                        (imovel_id, animal_id, ordem_parto, data_parto,
                         duracao_lactacao_dias, producao_total_litros,
                         producao_305d_litros, producao_acumulada_gordura,
                         producao_acumulada_proteina, escore_corporal,
                         raca_registro, ccs_media, data_encerramento,
                         causa_encerramento, fonte)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'gisleite')
                    """,
                    (
                        data.imovel_id, item.animal_id, item.ordem_parto, item.data_parto,
                        item.duracao_lactacao_dias, item.producao_total_litros,
                        item.producao_305d_litros, item.producao_acumulada_gordura,
                        item.producao_acumulada_proteina, item.escore_corporal,
                        item.raca_registro, item.ccs_media, item.data_encerramento,
                        item.causa_encerramento,
                    ),
                )
                cur.execute("RELEASE SAVEPOINT sp_item")
                criados += 1
            except psycopg2.errors.UniqueViolation:
                cur.execute("ROLLBACK TO SAVEPOINT sp_item")
                duplicados.append({
                    "animal_id": item.animal_id,
                    "data_parto": str(item.data_parto),
                    "motivo": "Ja existe um registro de lactacao para este animal com essa data de parto.",
                })
                continue
            except Exception as e:
                cur.execute("ROLLBACK TO SAVEPOINT sp_item")
                erros.append({"animal_id": item.animal_id, "data_parto": str(item.data_parto), "erro": str(e)})
                continue
        conn.commit()
        return {
            "ok": True,
            "criados": criados,
            "duplicados": duplicados,
            "erros": erros,
            "total": len(data.itens),
        }
    finally:
        conn.close()


@router.get("/leiteiro/ordenha/resumo/{imovel_id}")
def resumo_ordenha(imovel_id: int, meses: int = Query(6, ge=1, le=24)):
    """Resumo mensal de produção de leite por ordenha."""
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT
              DATE_TRUNC('month', data)::date AS mes,
              SUM(volume_l) AS total_l,
              AVG(volume_l) AS media_dia_l,
              SUM(valor_total) AS receita,
              AVG(gordura_pct) AS gordura_media,
              AVG(proteina_pct) AS proteina_media,
              COUNT(*) AS registros
            FROM bovino_ordenha
            WHERE imovel_id=%s AND data >= CURRENT_DATE - (%s * 30)
            GROUP BY 1 ORDER BY 1 DESC
        """, (imovel_id, meses))
        return list(cur.fetchall())
    finally:
        conn.close()

# ── IATF ─────────────────────────────────────────────────────
@router.post("/leiteiro/iatf")
def registrar_iatf(data: IatfIn):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO bovino_protocolo_iatf
              (imovel_id, femea_id, lote_id, protocolo, data_inicio, data_iatf,
               touro_id, semen_touro, tecnico, resultado, data_diagnostico, observacoes)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (
            data.imovel_id, data.femea_id, data.lote_id, data.protocolo,
            data.data_inicio, data.data_iatf, data.touro_id, data.semen_touro,
            data.tecnico, data.resultado, data.data_diagnostico, data.observacoes
        ))
        row = dict(cur.fetchone())
        conn.commit()
        return {"ok": True, **row}
    finally:
        conn.close()

@router.get("/leiteiro/iatf/{imovel_id}")
def listar_iatf(imovel_id: int):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT i.*, a.brinco, a.nome AS nome_femea
            FROM bovino_protocolo_iatf i
            JOIN bovino_animais a ON a.id = i.femea_id
            WHERE i.imovel_id=%s
            ORDER BY i.data_inicio DESC
        """, (imovel_id,))
        return list(cur.fetchall())
    finally:
        conn.close()

@router.patch("/leiteiro/iatf/{iatf_id}/resultado")
def atualizar_resultado_iatf(iatf_id: int, resultado: str, data_diagnostico: Optional[date] = None):
    if resultado not in ("aguardando", "positivo", "negativo", "aborto"):
        raise HTTPException(400, "resultado inválido")
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            UPDATE bovino_protocolo_iatf
            SET resultado=%s, data_diagnostico=COALESCE(%s, data_diagnostico)
            WHERE id=%s RETURNING id
        """, (resultado, data_diagnostico, iatf_id))
        if not cur.fetchone():
            raise HTTPException(404, "Registro não encontrado")
        conn.commit()
        return {"ok": True}
    finally:
        conn.close()

# ── Dieta de Transição ────────────────────────────────────────
@router.post("/leiteiro/dieta-transicao")
def registrar_dieta(data: DietaTransicaoIn):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO bovino_dieta_transicao
              (imovel_id, animal_id, fase, data_inicio, data_fim,
               dieta_descricao, volumoso_kg_dia, concentrado_kg_dia,
               suplemento, responsavel, observacoes)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (
            data.imovel_id, data.animal_id, data.fase, data.data_inicio,
            data.data_fim, data.dieta_descricao, data.volumoso_kg_dia,
            data.concentrado_kg_dia, data.suplemento, data.responsavel, data.observacoes
        ))
        row = dict(cur.fetchone())
        conn.commit()
        return {"ok": True, **row}
    finally:
        conn.close()

@router.get("/leiteiro/dieta-transicao/{imovel_id}")
def listar_dietas(imovel_id: int):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT d.*, a.brinco, a.nome AS nome_animal
            FROM bovino_dieta_transicao d
            JOIN bovino_animais a ON a.id = d.animal_id
            WHERE d.imovel_id=%s
            ORDER BY d.data_inicio DESC
        """, (imovel_id,))
        return list(cur.fetchall())
    finally:
        conn.close()

# ══════════════════════════════════════════════════════════════
# ENDPOINTS — Gado de Corte: Confinamento
# ══════════════════════════════════════════════════════════════
@router.post("/corte/confinamento")
def iniciar_confinamento(data: ConfinamentoIn):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO bovino_confinamento
              (imovel_id, lote_id, data_entrada, data_saida_prev,
               peso_entrada_kg, dieta, custo_diario_cab, objetivo)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (
            data.imovel_id, data.lote_id, data.data_entrada, data.data_saida_prev,
            data.peso_entrada_kg, data.dieta, data.custo_diario_cab, data.objetivo
        ))
        row = dict(cur.fetchone())
        conn.commit()
        return {"ok": True, **row}
    finally:
        conn.close()

@router.get("/corte/confinamento/{imovel_id}")
def listar_confinamentos(imovel_id: int, status: Optional[str] = None):
    conn = get_db()
    try:
        cur = conn.cursor()
        filtro = "AND cf.status=%s" if status else ""
        params = (imovel_id, status) if status else (imovel_id,)
        cur.execute(f"""
            SELECT cf.*, bl.nome AS nome_lote,
                   (SELECT COUNT(*) FROM bovino_animais a WHERE a.lote_id = cf.lote_id AND a.status='ativo') AS qtd_animais
            FROM bovino_confinamento cf
            JOIN bovino_lotes bl ON bl.id = cf.lote_id
            WHERE cf.imovel_id=%s {filtro}
            ORDER BY cf.data_entrada DESC
        """, params)
        return list(cur.fetchall())
    finally:
        conn.close()

@router.patch("/corte/confinamento/{confinamento_id}/fechar")
def fechar_confinamento(confinamento_id: int, data: ConfinamentoFechamentoIn):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            UPDATE bovino_confinamento
            SET data_saida_real=%s, peso_saida_kg=%s, status=%s
            WHERE id=%s RETURNING id, gmd_kg
        """, (data.data_saida_real, data.peso_saida_kg, data.status, confinamento_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Confinamento não encontrado")
        conn.commit()
        return {"ok": True, "id": row["id"], "gmd_kg": row["gmd_kg"]}
    finally:
        conn.close()

# ── Classificação de Carcaça ──────────────────────────────────
@router.post("/corte/classificacao-carcaca")
def registrar_classificacao(data: ClassificacaoCarcacaIn):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO bovino_classificacao_carcaca
              (imovel_id, animal_id, abate_id, data, frigorifico,
               maturidade, acabamento, conformacao, peso_carcaca_kg,
               rendimento_pct, preco_arroba, valor_total, nota_fiscal, observacoes)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (
            data.imovel_id, data.animal_id, data.abate_id, data.data,
            data.frigorifico, data.maturidade, data.acabamento, data.conformacao,
            data.peso_carcaca_kg, data.rendimento_pct, data.preco_arroba,
            data.valor_total, data.nota_fiscal, data.observacoes
        ))
        row = dict(cur.fetchone())
        conn.commit()
        return {"ok": True, **row}
    finally:
        conn.close()

@router.get("/corte/classificacao-carcaca/{imovel_id}")
def listar_classificacoes(imovel_id: int, dias: int = Query(90, ge=1, le=730)):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT cc.*, a.brinco, a.nome AS nome_animal
            FROM bovino_classificacao_carcaca cc
            JOIN bovino_animais a ON a.id = cc.animal_id
            WHERE cc.imovel_id=%s AND cc.data >= CURRENT_DATE - %s
            ORDER BY cc.data DESC
        """, (imovel_id, dias))
        return list(cur.fetchall())
    finally:
        conn.close()

# ── Custo de Produção ─────────────────────────────────────────
@router.post("/corte/custo-producao")
def registrar_custo(data: CustoProducaoIn):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO bovino_custo_producao
              (imovel_id, lote_id, confinamento_id, periodo_inicio, periodo_fim,
               categoria, descricao, valor)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (
            data.imovel_id, data.lote_id, data.confinamento_id,
            data.periodo_inicio, data.periodo_fim,
            data.categoria, data.descricao, data.valor
        ))
        row = dict(cur.fetchone())
        conn.commit()
        return {"ok": True, **row}
    finally:
        conn.close()

@router.get("/corte/custo-producao/{imovel_id}")
def listar_custos(imovel_id: int, lote_id: Optional[int] = None):
    conn = get_db()
    try:
        cur = conn.cursor()
        filtro = "AND lote_id=%s" if lote_id else ""
        params = (imovel_id, lote_id) if lote_id else (imovel_id,)
        cur.execute(f"""
            SELECT cp.*, bl.nome AS nome_lote
            FROM bovino_custo_producao cp
            LEFT JOIN bovino_lotes bl ON bl.id = cp.lote_id
            WHERE cp.imovel_id=%s {filtro}
            ORDER BY cp.periodo_inicio DESC
        """, params)
        return list(cur.fetchall())
    finally:
        conn.close()

@router.get("/corte/custo-producao/resumo/{imovel_id}")
def resumo_custos(imovel_id: int, lote_id: Optional[int] = None):
    """Resumo de custos por categoria para o imóvel/lote."""
    conn = get_db()
    try:
        cur = conn.cursor()
        filtro = "AND lote_id=%s" if lote_id else ""
        params = (imovel_id, lote_id) if lote_id else (imovel_id,)
        cur.execute(f"""
            SELECT categoria, SUM(valor) AS total, COUNT(*) AS registros
            FROM bovino_custo_producao
            WHERE imovel_id=%s {filtro}
            GROUP BY categoria ORDER BY total DESC
        """, params)
        return list(cur.fetchall())
    finally:
        conn.close()

# ══════════════════════════════════════════════════════════════
# DASHBOARD UNIFICADO (leiteiro + corte)
# ══════════════════════════════════════════════════════════════
@router.get("/dashboard-v2/{imovel_id}")
def dashboard_v2(imovel_id: int):
    """Dashboard unificado com dados condicionais por tipo_bovino."""
    conn = get_db()
    try:
        cur = conn.cursor()
        # tipo do imóvel
        cur.execute("SELECT COALESCE(tipo_bovino,'corte') AS tipo FROM imoveis_rurais WHERE id=%s", (imovel_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Imóvel não encontrado")
        tipo = row["tipo"]

        # plantel
        cur.execute("""
            SELECT
              COUNT(*) FILTER (WHERE status='ativo') AS total,
              COUNT(*) FILTER (WHERE aptidao_manejo='leite' AND status='ativo') AS leite,
              COUNT(*) FILTER (WHERE aptidao_manejo='corte' AND status='ativo') AS corte
            FROM bovino_animais WHERE imovel_id=%s
        """, (imovel_id,))
        plantel = dict(cur.fetchone())

        result = {"tipo_bovino": tipo, "plantel": plantel}

        # dados leiteiro
        if tipo in ("leite", "misto"):
            cur.execute("""
                SELECT
                  COALESCE(SUM(volume_l),0) AS total_l_30d,
                  COALESCE(AVG(volume_l),0) AS media_dia_l,
                  COALESCE(SUM(valor_total),0) AS receita_30d
                FROM bovino_ordenha
                WHERE imovel_id=%s AND data >= CURRENT_DATE - 30
            """, (imovel_id,))
            result["leiteiro"] = dict(cur.fetchone())

            cur.execute("""
                SELECT COUNT(*) AS iatf_aguardando
                FROM bovino_protocolo_iatf
                WHERE imovel_id=%s AND resultado='aguardando'
            """, (imovel_id,))
            result["leiteiro"]["iatf_aguardando"] = cur.fetchone()["iatf_aguardando"]

        # dados corte
        if tipo in ("corte", "misto"):
            cur.execute("""
                SELECT
                  COUNT(*) FILTER (WHERE status='ativo') AS confinamentos_ativos,
                  COALESCE(SUM(custo_diario_cab) FILTER (WHERE status='ativo'), 0) AS custo_diario_total
                FROM bovino_confinamento WHERE imovel_id=%s
            """, (imovel_id,))
            result["corte"] = dict(cur.fetchone())

            cur.execute("""
                SELECT COALESCE(SUM(valor_total),0) AS receita_abates_90d
                FROM bovino_abates ab
                JOIN bovino_animais a ON a.id=ab.animal_id
                WHERE a.imovel_id=%s AND ab.data >= CURRENT_DATE - 90
            """, (imovel_id,))
            result["corte"]["receita_abates_90d"] = cur.fetchone()["receita_abates_90d"]

        return result
    finally:
        conn.close()

# ── WEBHOOK WHATSAPP/TELEGRAM (IA) ──────────────────────────────────────────

class WhatsAppMensagemBovino(BaseModel):
    telefone: str
    tipo_midia: str = "texto"
    conteudo: str
    imovel_id: Optional[int] = None


def _buscar_animal_bovino(cur, brinco: Optional[str], imovel_id: int) -> Optional[dict]:
    if not brinco:
        return None
    cur.execute(
        "SELECT id, brinco FROM bovino_animais WHERE imovel_id=%s AND LOWER(brinco)=LOWER(%s)",
        (imovel_id, brinco)
    )
    row = cur.fetchone()
    return dict(row) if row else None


def _produtor_do_imovel_bovino(cur, imovel_id: int) -> Optional[int]:
    cur.execute("SELECT produtor_id FROM imoveis_rurais WHERE id = %s", (imovel_id,))
    row = cur.fetchone()
    return row["produtor_id"] if row else None


def _criar_lancamento_lcdpr_bovino(conn, produtor_id, data, tipo: str, valor: float,
                                    descricao: str, origem: str = "whatsapp_bovino"):
    """Cria lançamento LCDPR em conexão própria (mesmo padrão de piscicultura.py / ovino.py)."""
    tipo_lancamento = "Receita" if tipo == "receita" else "Despesa"
    lcdpr_conn = None
    try:
        lcdpr_conn = get_db()
        with lcdpr_conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT id FROM subcontas WHERE LOWER(tipo) = LOWER(%s) LIMIT 1", (tipo_lancamento,))
            sub = cur.fetchone()
            subconta_id = sub["id"] if sub else None
            cur.execute("""
                INSERT INTO lancamentos (produtor_id, subconta_id, valor, data, origem)
                VALUES (%s, %s, %s, %s, %s) RETURNING id
            """, (produtor_id, subconta_id, float(valor), data, origem))
            row = cur.fetchone()
            lcdpr_conn.commit()
            return row["id"] if row else None
    except Exception as e:
        logger.error("[BOVINO] Erro ao criar lançamento LCDPR: %s", e)
        if lcdpr_conn:
            lcdpr_conn.rollback()
        return None
    finally:
        if lcdpr_conn:
            lcdpr_conn.close()


@router.post("/webhook-whatsapp")
def webhook_whatsapp_bovino(payload: WhatsAppMensagemBovino):
    from app.services.bovino_ia import classificar_mensagem_sync

    classificacao = classificar_mensagem_sync(texto=payload.conteudo, imovel_id=payload.imovel_id)
    intent = classificacao["intent"]
    entidades = classificacao["entidades"]
    confianca = classificacao["confianca"]
    resumo = classificacao["resumo"]
    evento_id = None
    evento_tab = None
    status_log = "processado"
    erro_msg = None

    conn = get_db()
    try:
        cur = conn.cursor()

        if confianca >= 0.5 and payload.imovel_id:

            if intent == "pesagem":
                animal = _buscar_animal_bovino(cur, entidades.get("brinco"), payload.imovel_id)
                if animal:
                    cur.execute("""
                        INSERT INTO bovino_pesagens (animal_id, data, peso_kg, motivo)
                        VALUES (%s,%s,%s,%s) RETURNING id
                    """, (animal["id"], entidades.get("data_evento"), entidades.get("peso_kg"),
                          entidades.get("motivo", "rotina")))
                    evento_id = cur.fetchone()["id"]
                    evento_tab = "bovino_pesagens"
                else:
                    status_log = "pendente"
                    resumo = f"Não encontrei o animal {entidades.get('brinco')}. Confira o brinco."

            elif intent == "producao_leite":
                animal = _buscar_animal_bovino(cur, entidades.get("brinco"), payload.imovel_id)
                cur.execute("""
                    INSERT INTO bovino_producao_leite
                        (imovel_id, animal_id, data, turno, volume_l, destinacao)
                    VALUES (%s,%s,%s,%s,%s,%s) RETURNING id
                """, (payload.imovel_id, animal["id"] if animal else None,
                      entidades.get("data_evento"), entidades.get("turno", "total"),
                      entidades.get("volume_l"), entidades.get("destinacao", "venda")))
                evento_id = cur.fetchone()["id"]
                evento_tab = "bovino_producao_leite"

            elif intent in ("vacinacao", "vermifugacao", "tratamento"):
                animal = _buscar_animal_bovino(cur, entidades.get("brinco"), payload.imovel_id)
                cur.execute("""
                    INSERT INTO bovino_sanitario
                        (animal_id, tipo, produto, dose_ml, data_aplicacao, observacoes)
                    VALUES (%s,%s,%s,%s,%s,%s) RETURNING id
                """, (animal["id"] if animal else None, intent, entidades.get("produto"),
                      entidades.get("dose_ml"), entidades.get("data_evento"),
                      entidades.get("diagnostico")))
                evento_id = cur.fetchone()["id"]
                evento_tab = "bovino_sanitario"

            elif intent == "cobertura":
                femea = _buscar_animal_bovino(cur, entidades.get("brinco_femea"), payload.imovel_id)
                touro = _buscar_animal_bovino(cur, entidades.get("brinco_touro"), payload.imovel_id)
                if femea:
                    cur.execute("""
                        INSERT INTO bovino_reproducao (femea_id, touro_id, metodo, data_cobertura)
                        VALUES (%s,%s,%s,%s) RETURNING id
                    """, (femea["id"], touro["id"] if touro else None,
                          entidades.get("metodo", "monta_natural"), entidades.get("data_evento")))
                    evento_id = cur.fetchone()["id"]
                    evento_tab = "bovino_reproducao"
                else:
                    status_log = "pendente"
                    resumo = f"Não encontrei a fêmea {entidades.get('brinco_femea')}. Confira o brinco."

            elif intent == "parto":
                femea = _buscar_animal_bovino(cur, entidades.get("brinco_matriz"), payload.imovel_id)
                if femea:
                    cur.execute("""
                        UPDATE bovino_reproducao SET data_parto_real = %s
                        WHERE femea_id = %s AND resultado = 'positivo' AND data_parto_real IS NULL
                        RETURNING id
                    """, (entidades.get("data_evento"), femea["id"]))
                    row = cur.fetchone()
                    if row:
                        evento_id = row["id"]
                        evento_tab = "bovino_reproducao"
                    else:
                        status_log = "pendente"
                        resumo = "Parto registrado no resumo, mas não achei uma cobertura em aberto para essa matriz — confira manualmente."
                else:
                    status_log = "pendente"
                    resumo = f"Não encontrei a matriz {entidades.get('brinco_matriz')}. Confira o brinco."

            elif intent == "abate":
                animal = _buscar_animal_bovino(cur, entidades.get("brinco"), payload.imovel_id)
                if animal:
                    cur.execute("""
                        INSERT INTO bovino_abates
                            (animal_id, data, tipo, peso_vivo_kg, peso_carcaca_kg,
                             preco_arroba, valor_total, comprador)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
                    """, (animal["id"], entidades.get("data_evento"), "abate_frigorif",
                          entidades.get("peso_vivo_kg"), entidades.get("peso_carcaca_kg"),
                          entidades.get("preco_arroba"), entidades.get("valor_total"),
                          entidades.get("comprador")))
                    evento_id = cur.fetchone()["id"]
                    evento_tab = "bovino_abates"
                    cur.execute("UPDATE bovino_animais SET status='abatido', updated_at=NOW() WHERE id=%s", (animal["id"],))
                else:
                    status_log = "pendente"
                    resumo = f"Não encontrei o animal {entidades.get('brinco')}. Confira o brinco."

            elif intent == "cadastro":
                try:
                    cur.execute("SELECT id FROM especie WHERE codigo = 'BOVINO'")
                    especie_row = cur.fetchone()
                    especie_id = especie_row["id"] if especie_row else None
                    cur.execute("""
                        INSERT INTO bovino_animais
                            (imovel_id, especie_id, brinco, sexo, categoria, aptidao_manejo,
                             data_nascimento, origem)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,'nascimento') RETURNING id
                    """, (payload.imovel_id, especie_id, entidades.get("brinco"),
                          entidades.get("sexo", "F"), entidades.get("categoria", "bezerro"),
                          "corte", entidades.get("data_nascimento")))
                    evento_id = cur.fetchone()["id"]
                    evento_tab = "bovino_animais"
                except psycopg2.errors.UniqueViolation:
                    conn.rollback()
                    resumo = f"Animal {entidades.get('brinco')} já cadastrado."

            elif intent == "morte":
                animal = _buscar_animal_bovino(cur, entidades.get("brinco"), payload.imovel_id)
                if animal:
                    cur.execute("UPDATE bovino_animais SET status='morto', updated_at=NOW() WHERE id=%s RETURNING id",
                                (animal["id"],))
                    evento_id = cur.fetchone()["id"]
                    evento_tab = "bovino_animais"
                else:
                    status_log = "pendente"
                    resumo = f"Não encontrei o animal {entidades.get('brinco')}. Confira o brinco."

            elif intent == "compra":
                produtor_id = _produtor_do_imovel_bovino(cur, payload.imovel_id)
                valor_total = entidades.get("valor_total")
                qtd = entidades.get("quantidade")
                if valor_total:
                    lanc_id = _criar_lancamento_lcdpr_bovino(
                        conn, produtor_id, entidades.get("data_evento"), "despesa",
                        valor_total, f"Compra de {qtd or '?'} bovino(s)"
                                     + (f" — {entidades['raca']}" if entidades.get("raca") else ""),
                    )
                    if lanc_id:
                        evento_id = lanc_id
                        evento_tab = "lancamentos"
                        resumo = (
                            f"✅ Compra registrada: {qtd or '?'} animal(is) por "
                            f"R$ {float(valor_total):,.2f}. Cadastre os animais individualmente "
                            f"(com brinco) quando possível para o rebanho ficar completo."
                        )
                    else:
                        status_log = "erro"
                        resumo = "Entendi a compra, mas não consegui gravar o lançamento financeiro. Confira manualmente."
                else:
                    status_log = "pendente"
                    resumo = "Entendi que foi uma compra, mas não identifiquei o valor. Pode informar o valor total?"

            elif intent == "venda":
                produtor_id = _produtor_do_imovel_bovino(cur, payload.imovel_id)
                valor_total = entidades.get("valor_total")
                qtd = entidades.get("quantidade")
                if valor_total:
                    lanc_id = _criar_lancamento_lcdpr_bovino(
                        conn, produtor_id, entidades.get("data_evento"), "receita",
                        valor_total, f"Venda de {qtd or '?'} bovino(s)"
                                     + (f" — brinco {entidades['brinco']}" if entidades.get("brinco") else ""),
                    )
                    if lanc_id:
                        evento_id = lanc_id
                        evento_tab = "lancamentos"
                        resumo = f"✅ Venda registrada: {qtd or '?'} animal(is) por R$ {float(valor_total):,.2f}."
                        brinco_vendido = entidades.get("brinco")
                        if brinco_vendido:
                            animal = _buscar_animal_bovino(cur, brinco_vendido, payload.imovel_id)
                            if animal:
                                cur.execute("UPDATE bovino_animais SET status='vendido', updated_at=NOW() WHERE id=%s", (animal["id"],))
                    else:
                        status_log = "erro"
                        resumo = "Entendi a venda, mas não consegui gravar o lançamento financeiro. Confira manualmente."
                else:
                    status_log = "pendente"
                    resumo = "Entendi que foi uma venda, mas não identifiquei o valor. Pode informar o valor total?"

            else:
                status_log = "ignorado"

        elif confianca < 0.5:
            status_log = "pendente"
            resumo = "Não entendi bem. Pode repetir com mais detalhes?"
        else:
            status_log = "ignorado"

        conn.commit()

    except Exception as e:
        conn.rollback()
        status_log = "erro"
        erro_msg = str(e)
        resumo = "Erro ao salvar. Tente novamente."
        logger.error("webhook_bovino erro: %s", e, exc_info=True)

    # Log da mensagem
    try:
        cur2 = conn.cursor()
        cur2.execute("""
            INSERT INTO bovino_whatsapp_log
                (telefone, tipo_midia, conteudo_raw, intent_detectada,
                 entidades_json, status, evento_id, evento_tabela, erro_msg)
            VALUES (%s, %s, %s, %s, %s::jsonb, %s, %s, %s, %s)
        """, (payload.telefone, payload.tipo_midia, payload.conteudo[:2000],
              intent, json.dumps(entidades, default=str),
              status_log, evento_id, evento_tab, erro_msg))
        conn.commit()
    except Exception as e:
        logger.warning("Falha ao salvar log WhatsApp bovino: %s", e)
    finally:
        conn.close()

    return {
        "intent": intent,
        "confianca": confianca,
        "status": status_log,
        "resumo": resumo,
        "evento_id": evento_id,
        "evento_tabela": evento_tab,
    }




# ══════════════════════════════════════════════════════════════
# IMPORTAÇÃO DE GENEALOGIA (planilhas Gisleite/Embrapa)
# ══════════════════════════════════════════════════════════════

MAPA_RACAS = {
    "HOL": "Holandês",
    "NEL": "Nelore",
    "GUZ": "Guzerá",
    "GIR": "Gir Leiteiro",
    "GHL": "Girolando",   # composicao tipica "3/4 HOL + 1/4 GIR"
    "DES": "Sem raça definida",
    "CAR": "Caracu",
}
RACA_FALLBACK = "Mestiço"  # usado para codigos de cruzamento sem correspondencia direta


def _idade_em_meses(data_nasc, hoje):
    if not data_nasc:
        return None
    return (hoje.year - data_nasc.year) * 12 + (hoje.month - data_nasc.month)


def _categoria_por_idade(sexo: str, idade_meses):
    if idade_meses is None:
        return "indefinida"
    if sexo == "M":
        return "touro" if idade_meses >= 24 else "bezerro"
    return "vaca" if idade_meses >= 24 else ("novilha" if idade_meses >= 12 else "bezerra")


def _parse_identificador(linha: dict) -> str:
    ident = str(linha.get("Identificador Animal", "")).strip()
    if ident.endswith(".0"):
        ident = ident[:-2]
    return ident


def _parse_data_nascimento(linha: dict, _date):
    try:
        dia = int(float(linha.get("Data Nascimento Dia")))
        mes = int(float(linha.get("Data Nascimento Mes")))
        ano = int(float(linha.get("Data Nascimento Ano")))
        return _date(ano, mes, dia)
    except (TypeError, ValueError):
        return None


def resolver_brincos_duplicados(linhas: list, _date) -> tuple:
    """
    Brinco repetido dentro da própria planilha NÃO derruba a linha com erro.
    Critério: o animal mais ANTIGO quase certamente já saiu do rebanho
    (venda/morte) sem baixa registrada, e o brinco físico foi reaproveitado
    no animal mais novo. Então:
      - o mais RECENTE fica com o Identificador original;
      - o(s) mais antigo(s) recebem um brinco derivado (sufixo "-BXn") e uma
        observação explicando o motivo, para revisão manual (dar baixa/status
        depois, se for o caso).
    Duplicata literal (mesmo nome + mesma data) é linha repetida na planilha
    — mantém só uma, sem gerar erro nem conflito.
    Se faltar data de nascimento em algum lado do conflito, não dá pra decidir
    com segurança — a linha vai para "conflitos" em vez de erro ou palpite.

    Retorna (linhas_prontas, conflitos), onde cada dict em linhas_prontas
    pode ter a chave extra "_brinco_final" (quando renomeado) e "_obs_baixa".
    """
    grupos: dict[str, list[dict]] = {}
    for linha in linhas:
        ident = _parse_identificador(linha)
        grupos.setdefault(ident, []).append(linha)

    prontas: list[dict] = []
    conflitos: list[dict] = []

    for ident, grupo in grupos.items():
        if not ident or len(grupo) == 1:
            prontas.extend(grupo)
            continue

        for linha in grupo:
            linha["_data_nasc_parsed"] = _parse_data_nascimento(linha, _date)

        assinaturas = {
            (str(l.get("Nome Animal", "")).strip(), l["_data_nasc_parsed"])
            for l in grupo
        }
        if len(assinaturas) == 1:
            # duplicata literal — mantém só a primeira ocorrência
            prontas.append(grupo[0])
            continue

        if any(l["_data_nasc_parsed"] is None for l in grupo):
            for l in grupo:
                l["_conflito_motivo"] = (
                    f"Identificador '{ident}' duplicado na planilha e ao menos "
                    f"um dos registros está sem data de nascimento — não dá "
                    f"pra saber com segurança qual é o mais antigo. Revisar "
                    f"manualmente."
                )
            conflitos.extend(grupo)
            continue

        grupo_ordenado = sorted(grupo, key=lambda l: l["_data_nasc_parsed"])
        mais_novo = grupo_ordenado[-1]
        antigos = grupo_ordenado[:-1]
        prontas.append(mais_novo)

        for idx, antigo in enumerate(antigos, start=1):
            sufixo = f"-BX{idx}" if len(antigos) > 1 else "-BX"
            antigo["_brinco_final"] = f"{ident}{sufixo}"
            antigo["_obs_baixa"] = (
                f"Identificador original '{ident}' reaproveitado por animal mais "
                f"novo ('{mais_novo.get('Nome Animal')}', "
                f"nasc. {mais_novo['_data_nasc_parsed']}). Brinco renomeado "
                f"automaticamente para '{antigo['_brinco_final']}'. Provável "
                f"baixa (venda/morte) não registrada — revisar e marcar status "
                f"correto."
            )
            prontas.append(antigo)

    return prontas, conflitos


def _brinco_livre(base: str, em_uso: set) -> str:
    """Gera um sufixo -BXn livre, considerando tanto brincos já existentes no
    banco quanto os que já foram atribuídos nesta mesma importação (evita
    colisão entre renomeação por duplicidade-no-arquivo e por duplicidade-no-banco)."""
    candidato = f"{base}-BX"
    if candidato not in em_uso:
        em_uso.add(candidato)
        return candidato
    n = 2
    while f"{base}-BX{n}" in em_uso:
        n += 1
    candidato = f"{base}-BX{n}"
    em_uso.add(candidato)
    return candidato


def resolver_duplicata_com_banco(
    brinco: str, nome, data_nasc, existentes_por_brinco: dict, brincos_em_uso: set, cur, _date
):
    """
    Mesma regra do resolver_brincos_duplicados, mas comparando contra um
    animal JÁ CADASTRADO no banco (de uma importação anterior), não outra
    linha da mesma planilha.

    Retorna uma tupla (acao, brinco_final, observacao):
      - ("inserir", brinco, None)              -> segue pro INSERT normal
      - ("inserir", brinco_renomeado, obs)      -> segue pro INSERT, mas com
                                                    brinco alterado (o novo
                                                    animal é o mais antigo)
      - ("pular_identico", None, obs)           -> já existe idêntico, não
                                                    insere de novo (não é erro)
      - ("conflito", None, motivo)              -> não dá pra decidir, não
                                                    insere (vai pra revisão)
    Quando o existente no banco é o mais antigo, essa função também executa o
    UPDATE que renomeia o registro existente antes de liberar o brinco para o
    novo.
    """
    existente = existentes_por_brinco.get(brinco)
    if not existente:
        return "inserir", brinco, None

    nome_novo = (str(nome).strip() if nome else None)
    if existente["nome"] == nome_novo and existente["data_nascimento"] == data_nasc:
        return "pular_identico", None, (
            f"Brinco '{brinco}' já cadastrado com os mesmos dados (id={existente['id']}) — ignorado."
        )

    if data_nasc is None or existente["data_nascimento"] is None:
        return "conflito", None, (
            f"Brinco '{brinco}' já existe no rebanho (id={existente['id']}, "
            f"nome='{existente['nome']}'), com dados diferentes da planilha, e "
            f"não há data de nascimento suficiente para decidir qual é o mais "
            f"antigo. Revisar manualmente."
        )

    if existente["data_nascimento"] < data_nasc:
        # o registro do banco é o mais antigo -> renomeia ELE, libera o brinco pro novo
        novo_brinco_existente = _brinco_livre(brinco, brincos_em_uso)
        cur.execute(
            "UPDATE bovino_animais SET brinco = %s, updated_at = NOW(), "
            "observacoes = COALESCE(observacoes || ' | ', '') || %s WHERE id = %s",
            (
                novo_brinco_existente,
                f"Brinco original '{brinco}' reaproveitado por animal mais novo "
                f"importado em {_date.today().isoformat()}. Renomeado automaticamente "
                f"para '{novo_brinco_existente}'. Provável baixa (venda/morte) não "
                f"registrada — revisar e marcar status correto.",
                existente["id"],
            ),
        )
        existentes_por_brinco.pop(brinco, None)
        return "inserir", brinco, None
    else:
        # a linha da planilha e a mais antiga -> ela e' quem recebe o sufixo
        brinco_renomeado = _brinco_livre(brinco, brincos_em_uso)
        obs = (
            f"Brinco original '{brinco}' já em uso por animal mais novo já "
            f"cadastrado (id={existente['id']}, nome='{existente['nome']}'). "
            f"Renomeado automaticamente para '{brinco_renomeado}'. Provável "
            f"baixa (venda/morte) não registrada — revisar e marcar status correto."
        )
        return "inserir", brinco_renomeado, obs


def _ler_planilha_generica(conteudo: bytes, nome_arquivo: str):
    """Tenta abrir como xlsx real, depois xls binario real, depois HTML
    disfarçado de xls (comum em exports de sistemas como Gisleite)."""
    import io
    # 1. xlsx real
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(conteudo), data_only=True)
        ws = wb.active
        linhas = list(ws.iter_rows(values_only=True))
        cabecalho = [str(c or "").strip() for c in linhas[0]]
        return [dict(zip(cabecalho, row)) for row in linhas[1:] if any(str(c or "").strip() for c in row)]
    except Exception:
        pass
    # 2. xls binario real
    try:
        import xlrd
        wb = xlrd.open_workbook(file_contents=conteudo)
        ws = wb.sheet_by_index(0)
        cabecalho = [str(c or "").strip() for c in ws.row_values(0)]
        linhas = []
        for i in range(1, ws.nrows):
            row = ws.row_values(i)
            if any(str(c).strip() for c in row):
                linhas.append(dict(zip(cabecalho, row)))
        return linhas
    except Exception:
        pass
    # 3. HTML disfarçado de xls
    import pandas as pd
    tabelas = pd.read_html(io.BytesIO(conteudo))
    df = tabelas[0]
    return df.to_dict(orient="records")


@router.post("/importar-genealogia")
async def importar_genealogia(
    arquivo: UploadFile = File(...),
    imovel_id: int = Form(...),
):
    from datetime import date as _date

    conteudo = await arquivo.read()
    try:
        linhas = _ler_planilha_generica(conteudo, arquivo.filename or "")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Não foi possível ler o arquivo: {e}")

    if not linhas:
        return {"criados": 0, "erros": 0, "total": 0, "mensagem": "Nenhuma linha de dados encontrada."}

    linhas, conflitos_duplicados = resolver_brincos_duplicados(linhas, _date)

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id FROM especie WHERE codigo = 'BOVINO'")
    especie_id = cur.fetchone()["id"]

    cur.execute("SELECT id, nome, aptidao FROM bovino_racas")
    racas_cadastradas = {r["nome"]: r for r in cur.fetchall()}

    cur.execute(
        "SELECT id, brinco, nome, data_nascimento FROM bovino_animais WHERE imovel_id = %s",
        (imovel_id,),
    )
    existentes_por_brinco = {r["brinco"]: r for r in cur.fetchall()}
    brincos_em_uso = set(existentes_por_brinco.keys())

    hoje = _date.today()
    criados = 0
    ja_existentes = 0
    renomeados_banco = 0
    erros_lista = []
    racas_nao_mapeadas: dict[str, int] = {}
    # brinco ORIGINAL (Identificador Animal) -> id recem-criado, para linkar
    # mae/pai na 2a passada — usa o identificador original mesmo quando o
    # brinco gravado no banco foi renomeado por duplicidade (-BXn), porque é
    # esse valor original que aparece em "Registro Pai"/"Registro Mae" de
    # outros animais do mesmo arquivo.
    mapa_brinco_para_id: dict[str, int] = {}
    pendencias_genealogia = []  # (novo_id, registro_pai, registro_mae)

    for i, linha in enumerate(linhas, start=2):
        try:
            brinco_original = _parse_identificador(linha)
            if not brinco_original:
                erros_lista.append(f"Linha {i}: identificador do animal ausente")
                continue
            brinco = linha.get("_brinco_final") or brinco_original

            nome = str(linha.get("Nome Animal", "") or "").strip() or None
            sexo_raw = str(linha.get("Sexo Animal", "") or "").strip().upper()
            sexo = "M" if sexo_raw.startswith("M") else "F"

            codigo_raca = str(linha.get("Raca Animal", "") or "").strip().upper()
            nome_raca = MAPA_RACAS.get(codigo_raca)
            if not nome_raca:
                racas_nao_mapeadas[codigo_raca] = racas_nao_mapeadas.get(codigo_raca, 0) + 1
                nome_raca = RACA_FALLBACK
            raca_info = racas_cadastradas.get(nome_raca) or racas_cadastradas.get(RACA_FALLBACK)
            raca_id = raca_info["id"] if raca_info else None
            aptidao_manejo = raca_info["aptidao"] if raca_info and raca_info["aptidao"] in ("leite", "corte") else "leite"

            data_nasc = linha.get("_data_nasc_parsed", "AUSENTE")
            if data_nasc == "AUSENTE":
                data_nasc = _parse_data_nascimento(linha, _date)

            acao, brinco_resolvido, obs_banco = resolver_duplicata_com_banco(
                brinco, nome, data_nasc, existentes_por_brinco, brincos_em_uso, cur, _date
            )
            if acao == "pular_identico":
                ja_existentes += 1
                continue
            if acao == "conflito":
                linha["_conflito_motivo"] = obs_banco
                conflitos_duplicados.append(linha)
                continue
            if brinco_resolvido != brinco:
                renomeados_banco += 1
            brinco = brinco_resolvido

            idade_meses = _idade_em_meses(data_nasc, hoje)
            categoria = _categoria_por_idade(sexo, idade_meses)

            registro_pai = str(linha.get("Registro Pai", "") or "").strip()
            nome_pai = str(linha.get("Nome Pai", "") or "").strip()
            registro_mae = str(linha.get("Registro Mae", "") or "").strip()
            nome_mae = str(linha.get("Nome Mae", "") or "").strip()
            for sufixo in (".0",):
                if registro_pai.endswith(sufixo):
                    registro_pai = registro_pai[:-2]
                if registro_mae.endswith(sufixo):
                    registro_mae = registro_mae[:-2]

            obs_partes = []
            if nome_pai and nome_pai.lower() != "desconhecido":
                obs_partes.append(f"Pai: {nome_pai} (registro {registro_pai})")
            if nome_mae and nome_mae.lower() != "desconhecido":
                obs_partes.append(f"Mãe: {nome_mae} (registro {registro_mae})")
            if linha.get("_obs_baixa"):
                obs_partes.append(linha["_obs_baixa"])
            if obs_banco:
                obs_partes.append(obs_banco)
            observacoes = " | ".join(obs_partes) or None

            cur.execute("""
                INSERT INTO bovino_animais
                (imovel_id, especie_id, brinco, nome, raca_id, sexo, aptidao_manejo,
                 categoria, data_nascimento, mae_id, pai_id, data_entrada, origem, observacoes)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,NULL,NULL,%s,'importacao',%s)
                RETURNING id
            """, (imovel_id, especie_id, brinco, nome, raca_id, sexo, aptidao_manejo,
                  categoria, data_nasc, hoje, observacoes))
            novo_id = cur.fetchone()["id"]
            mapa_brinco_para_id[brinco_original] = novo_id
            if registro_pai or registro_mae:
                pendencias_genealogia.append((novo_id, registro_pai, registro_mae))
            criados += 1
        except Exception as e:
            erros_lista.append(f"Linha {i}: {str(e)}")

    # 2a passada: linka mae_id/pai_id quando o registro bate com um brinco importado nesta mesma leva
    linkados = 0
    for novo_id, registro_pai, registro_mae in pendencias_genealogia:
        pai_id = mapa_brinco_para_id.get(registro_pai)
        mae_id = mapa_brinco_para_id.get(registro_mae)
        if pai_id or mae_id:
            cur.execute(
                "UPDATE bovino_animais SET pai_id = COALESCE(%s, pai_id), mae_id = COALESCE(%s, mae_id) WHERE id = %s",
                (pai_id, mae_id, novo_id),
            )
            linkados += 1

    conn.commit()
    conn.close()

    mensagem_partes = []
    if erros_lista:
        mensagem_partes.append("; ".join(erros_lista[:10]))
    if conflitos_duplicados:
        idents_conflito = sorted({_parse_identificador(l) for l in conflitos_duplicados})
        mensagem_partes.append(
            "Identificador(es) duplicado(s) sem data de nascimento suficiente "
            "pra decidir automaticamente qual é o mais antigo — revisar "
            "manualmente: " + ", ".join(idents_conflito)
        )
    if ja_existentes:
        mensagem_partes.append(f"{ja_existentes} animal(is) já cadastrado(s) e idêntico(s) — ignorado(s).")
    if renomeados_banco:
        mensagem_partes.append(
            f"{renomeados_banco} brinco(s) renomeado(s) automaticamente por já "
            f"estarem em uso no rebanho (provável baixa não registrada)."
        )
    if racas_nao_mapeadas:
        mensagem_partes.append(
            "Raças sem mapeamento direto (usaram 'Mestiço', revisar manualmente): "
            + ", ".join(f"{k} ({v})" for k, v in racas_nao_mapeadas.items())
        )

    return {
        "criados": criados,
        "ja_existentes": ja_existentes,
        "renomeados_banco": renomeados_banco,
        "erros": len(erros_lista),
        "conflitos_duplicados": len(conflitos_duplicados),
        "total": len(linhas) + len(conflitos_duplicados),
        "genealogia_linkada": linkados,
        "mensagem": " | ".join(mensagem_partes) if mensagem_partes else None,
    }


print("BOVINO ROUTER LOADED OK")
